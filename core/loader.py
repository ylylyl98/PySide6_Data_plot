from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
import json
from pathlib import Path
from typing import Optional, Sequence, Tuple

import numpy as np

from core import processing_run as P


XLSX_Y_LABEL_DOPING = "Doping (V)"
XLSX_Y_LABEL_EFIELD = "Efield (V)"
XLSX_Y_LABEL_OPTIONS = (XLSX_Y_LABEL_DOPING, XLSX_Y_LABEL_EFIELD)
DAT_Y_AXIS_OPTIONS = ("Y", "Doping", "Electric field", "Gate voltage", "Custom")


def is_xlsx_map_file(file_name: str) -> bool:
    """Return True when a file name selects the precomputed XLSX map format."""
    return Path(str(file_name)).suffix.lower() == ".xlsx"


def resolve_xlsx_y_label(y_axis: str) -> str:
    """Map a y-axis request to a label-only XLSX axis label."""
    request = str(y_axis or "").strip().lower()
    if request in {"efield", "efield (v)"}:
        return XLSX_Y_LABEL_EFIELD
    return XLSX_Y_LABEL_DOPING


@dataclass
class DataCube:
    energy: np.ndarray
    gate: np.ndarray
    Z: np.ndarray
    gate_label: str
    title: str
    cbar_label: str
    gate_unit: str = ""
    y_axis_semantic: str = ""


def _dat_sidecar_candidates(path: Path) -> tuple[Path, ...]:
    return (
        path.with_suffix(".metadata.json"),
        Path(f"{path}.plotmeta.json"),
    )


def _load_dat_sidecar(path: Path) -> dict:
    for sidecar in _dat_sidecar_candidates(path):
        if not sidecar.is_file():
            continue
        try:
            value = json.loads(sidecar.read_text(encoding="utf-8"))
        except (OSError, UnicodeError, json.JSONDecodeError):
            return {}
        return value if isinstance(value, dict) else {}
    return {}


def resolve_dat_y_axis(choice: str, *, custom_label: str = "", custom_unit: str = "") -> tuple[str, str, str]:
    """Map a user-facing DAT Y-axis choice to label, unit, and semantic id."""
    selected = str(choice or "Y").strip()
    if selected == "Doping":
        return "Doping", str(custom_unit).strip(), "doping"
    if selected == "Electric field":
        return "Electric field", str(custom_unit).strip(), "electric_field"
    if selected == "Gate voltage":
        return "Gate voltage", str(custom_unit).strip(), "gate_voltage"
    if selected == "Custom":
        label = str(custom_label).strip() or "Y"
        return label, str(custom_unit).strip(), "custom"
    return "Y", "", "y"


def load_dat(path: str | Path) -> DataCube:
    """Load an Origin-friendly exported DAT matrix into the normal DataCube model."""
    dat_path = Path(path)
    if not dat_path.is_file():
        raise FileNotFoundError(f"DAT file not found: {dat_path}")
    try:
        lines = dat_path.read_text(encoding="utf-8-sig").splitlines()
    except (OSError, UnicodeError) as exc:
        raise ValueError(f"Could not read DAT file {dat_path}: {exc}") from exc
    content = [line for line in lines if line.strip() and not line.lstrip().startswith("#")]
    if not content:
        raise ValueError(f"DAT file {dat_path.name!r} contains no numeric table.")
    header = content[0].split("\t")
    if len(header) < 2:
        raise ValueError(f"DAT file {dat_path.name!r} must contain an X column and at least one Y column.")
    try:
        gate = np.asarray([float(value.strip()) for value in header[1:]], dtype=float)
    except ValueError as exc:
        raise ValueError(f"DAT file {dat_path.name!r} has non-numeric Y/gate headers.") from exc
    rows: list[list[float]] = []
    for row_number, line in enumerate(content[1:], start=2):
        fields = line.split("\t")
        if len(fields) != len(header):
            raise ValueError(
                f"DAT file {dat_path.name!r} row {row_number} has {len(fields)} columns; expected {len(header)}."
            )
        try:
            rows.append([float(value.strip()) for value in fields])
        except ValueError as exc:
            raise ValueError(f"DAT file {dat_path.name!r} has non-numeric data on row {row_number}.") from exc
    if not rows:
        raise ValueError(f"DAT file {dat_path.name!r} contains no data rows.")
    table = np.asarray(rows, dtype=float)
    energy = table[:, 0]
    z = table[:, 1:].T.copy()
    metadata = _load_dat_sidecar(dat_path)
    plot = metadata.get("plot", {}) if isinstance(metadata.get("plot", {}), dict) else {}
    if isinstance(plot.get("linear"), dict):
        plot = plot["linear"]
    processing = metadata.get("processing", {}) if isinstance(metadata.get("processing", {}), dict) else {}
    mode = str(metadata.get("operation", processing.get("mode", "")))
    sidecar_label = metadata.get("y_axis_label", processing.get("y_axis_label", plot.get("ylabel", "Y")))
    sidecar_unit = metadata.get("y_axis_unit", processing.get("y_axis_unit", plot.get("y_unit", "")))
    sidecar_semantic = metadata.get("y_axis_semantic", processing.get("y_axis_semantic", ""))
    if not sidecar_semantic:
        sidecar_semantic = {
            "doping": "doping",
            "electric field": "electric_field",
            "gate voltage": "gate_voltage",
        }.get(str(sidecar_label).strip().lower(), "" if str(sidecar_label).strip() in {"", "Y"} else "custom")
    cube = DataCube(
        energy=energy,
        gate=gate,
        Z=z,
        gate_label=str(sidecar_label or "Y"),
        title=str(plot.get("title", dat_path.stem)),
        cbar_label=str(plot.get("cbar_label", mode or "Imported DAT")),
        gate_unit=str(sidecar_unit or ""),
        y_axis_semantic=str(sidecar_semantic or ""),
    )
    cube.plot_metadata = plot
    cube.import_metadata = metadata
    return cube


def _validate_cube_arrays(energy: np.ndarray, gate: np.ndarray, Z: np.ndarray, *, context: str) -> None:
    e = np.asarray(energy).ravel()
    g = np.asarray(gate).ravel()
    z = np.asarray(Z)

    if e.size == 0 or g.size == 0 or z.size == 0:
        raise ValueError(f"{context}: empty energy/gate/Z data.")
    if z.ndim != 2:
        raise ValueError(f"{context}: Z must be 2D, got shape {z.shape}.")
    if z.shape not in {(g.size, e.size), (e.size, g.size)}:
        raise ValueError(
            f"{context}: Z shape {z.shape} does not match gate ({g.size}) x energy ({e.size})."
        )


def _csv_signature(user_folder: str, file_name: str) -> Tuple[int, int]:
    """Return cache key pieces from the root CSV mtime and size."""
    folder = Path(user_folder)
    if not folder.exists() or not folder.is_dir():
        raise FileNotFoundError(f"Folder does not exist: {user_folder}")

    csv_path = folder / Path(file_name).name
    if not csv_path.exists() or not csv_path.is_file():
        raise FileNotFoundError(f"CSV not found in folder root: {csv_path}")

    stt = csv_path.stat()
    return int(stt.st_mtime_ns), int(stt.st_size)


@lru_cache(maxsize=256)
def _peek_y_axis_options_cached(
    user_folder: str, file_name: str, csv_sig: Tuple[int, int]
) -> Tuple[tuple[str, ...], str]:
    del csv_sig

    # Preferred: use implementation that owns _load_canonical
    if hasattr(P, "peek_y_axis_options"):
        opts, default = P.peek_y_axis_options(user_folder, file_name)
    else:
        d = P._load_canonical(user_folder, file_name, y_axis="auto")  # type: ignore[attr-defined]
        opts = d.get("available_axes", ["Vbg", "Vtg"])
        default = d.get("default_axis", opts[0] if opts else "Vtg")

    if default not in opts and opts:
        default = opts[0]
    return tuple(str(o) for o in opts), str(default)


def peek_y_axis_options(user_folder: str, file_name: str) -> Tuple[list[str], str]:
    csv_sig = _csv_signature(user_folder, file_name)
    opts, default = _peek_y_axis_options_cached(user_folder, file_name, csv_sig)
    return list(opts), default


@lru_cache(maxsize=512)
def _load_pl_cached(
    user_folder: str, file_name: str, log_scale: bool, y_axis: str, csv_sig: Tuple[int, int]
) -> dict:
    del csv_sig

    return P.process_pl(
        user_folder=user_folder,
        file=file_name,
        y_axis=y_axis,
        plot_interactive=False,
        save_png=False,
        save_dat_file=False,
        move_original=False,
        pl_scales=("log" if log_scale else "linear",),
        open_both_interactive=False,
    )


def load_pl(user_folder: str, file_name: str, *, log_scale: bool = False, y_axis: str = "auto") -> DataCube:
    if is_xlsx_map_file(file_name):
        return load_xlsx_map(user_folder, file_name, y_label=resolve_xlsx_y_label(y_axis))

    csv_sig = _csv_signature(user_folder, file_name)
    effective_y_axis = P.resolve_shared_y_axis_request([file_name], y_axis)
    res = _load_pl_cached(user_folder, file_name, bool(log_scale), effective_y_axis, csv_sig)

    _validate_cube_arrays(res["energy"], res["gate_axis"], res["Z"], context="PL load")
    return DataCube(
        energy=np.asarray(res["energy"], dtype=float).copy(),
        gate=np.asarray(res["gate_axis"], dtype=float).copy(),
        Z=np.asarray(res["Z"], dtype=float).copy(),
        gate_label=res.get("gate_label", "Gate (V)"),
        title=res.get("title", file_name),
        cbar_label="PL (a.u.)",
    )


def _xlsx_signature(user_folder: str, file_name: str) -> Tuple[int, int]:
    """Return cache key pieces from the root XLSX mtime and size."""
    folder = Path(user_folder)
    if not folder.exists() or not folder.is_dir():
        raise FileNotFoundError(f"Folder does not exist: {user_folder}")

    xlsx_path = folder / Path(file_name).name
    if not xlsx_path.exists() or not xlsx_path.is_file():
        raise FileNotFoundError(f"XLSX not found in folder root: {xlsx_path}")

    stat = xlsx_path.stat()
    return int(stat.st_mtime_ns), int(stat.st_size)


def _xlsx_cell_value(cell, *, context: str) -> float:
    coordinate = getattr(cell, "coordinate", "?")
    if cell.data_type == "f":
        raise ValueError(f"{context}: formulas are not supported (cell {coordinate}).")
    value = cell.value
    if value is None or isinstance(value, bool):
        raise ValueError(f"{context}: empty cell at {coordinate}.")
    try:
        result = float(str(value).strip())
    except (TypeError, ValueError):
        raise ValueError(f"{context}: non-numeric value {value!r} at {coordinate}.")
    if not np.isfinite(result):
        raise ValueError(f"{context}: non-finite value at {coordinate}.")
    return result


@lru_cache(maxsize=32)
def _load_xlsx_map_cached(
    user_folder: str, file_name: str, signature: Tuple[int, int]
) -> Tuple[np.ndarray, np.ndarray, np.ndarray, str]:
    del signature

    from openpyxl import load_workbook

    path = Path(user_folder) / Path(file_name).name
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        sheet_names = workbook.sheetnames
        if "dR_R" in sheet_names:
            sheet = workbook["dR_R"]
        elif len(sheet_names) == 1:
            sheet = workbook[sheet_names[0]]
        else:
            raise ValueError(
                f"XLSX map {file_name!r} has {len(sheet_names)} sheets; "
                "expected a single 'dR_R' sheet."
            )
        rows = list(sheet.iter_rows(values_only=False))
    finally:
        workbook.close()

    if len(rows) < 2:
        raise ValueError(
            f"XLSX map {file_name!r} must contain a header row and at least one data row."
        )

    header = rows[0]
    if len(header) < 2:
        raise ValueError(
            f"XLSX map {file_name!r} header must include a corner cell and at least one y-value column."
        )

    if header[0].data_type == "f":
        raise ValueError(
            f"XLSX map {file_name!r}: formulas are not supported in the header corner "
            f"({header[0].coordinate})."
        )

    gate = np.asarray(
        [_xlsx_cell_value(cell, context=f"XLSX map {file_name!r} header") for cell in header[1:]],
        dtype=float,
    )
    if np.unique(gate).size != gate.size:
        raise ValueError(f"XLSX map {file_name!r} contains duplicate y-axis (doping) values.")

    energy_rows: list[float] = []
    z_rows: list[list[float]] = []
    expected = len(header)
    for row_index, row in enumerate(rows[1:], start=2):
        if len(row) != expected:
            raise ValueError(
                f"XLSX map {file_name!r}: row {row_index} has {len(row)} columns, expected {expected}."
            )
        energy_rows.append(
            _xlsx_cell_value(row[0], context=f"XLSX map {file_name!r} row {row_index} energy")
        )
        z_rows.append(
            [
                _xlsx_cell_value(cell, context=f"XLSX map {file_name!r} row {row_index}")
                for cell in row[1:]
            ]
        )

    energy = np.asarray(energy_rows, dtype=float)
    z = np.asarray(z_rows, dtype=float)

    order = np.argsort(energy, kind="stable")
    energy = energy[order]
    z = z[order, :]
    z = z.T

    return energy.copy(), gate.copy(), z.copy(), Path(file_name).name


def load_xlsx_map(
    user_folder: str, file_name: str, *, y_label: str = XLSX_Y_LABEL_DOPING
) -> DataCube:
    """Load a precomputed dR/R XLSX map into the standard DataCube contract."""
    raw_name = str(file_name)
    if Path(raw_name).name != raw_name:
        raise ValueError(
            f"XLSX map must reference a root-level file name, got {raw_name!r}"
        )
    if not is_xlsx_map_file(raw_name):
        raise ValueError(f"XLSX map must reference a .xlsx file, got {raw_name!r}")
    signature = _xlsx_signature(user_folder, file_name)
    energy, gate, z, title = _load_xlsx_map_cached(user_folder, Path(file_name).name, signature)
    _validate_cube_arrays(energy, gate, z, context="XLSX map load")
    return DataCube(
        energy=np.asarray(energy, dtype=float).copy(),
        gate=np.asarray(gate, dtype=float).copy(),
        Z=np.asarray(z, dtype=float).copy(),
        gate_label=resolve_xlsx_y_label(y_label),
        title=title,
        cbar_label="dR/R",
    )


def build_external_baseline(user_folder: str, files: Sequence[str], *, which: str = "last") -> dict:
    """
    Returns dict with keys: energy, I0

    which:
      - "first" : use first frame in each file
      - "last"  : use last frame in each file
      - "all"   : average ALL frames within each file, then average across files
    """
    if not files:
        raise ValueError("build_external_baseline: 'files' is empty.")

    w = (which or "last").strip().lower()
    alias = {
        "first": "first",
        "1st": "first",
        "start": "first",
        "last": "last",
        "end": "last",
        "all": "all",
        "avg": "all",
        "mean": "all",
        "all_frames": "all",
        "all frames": "all",
        "frames": "all",
    }
    w = alias.get(w, w)
    if w not in ("first", "last", "all"):
        raise ValueError(f"Unknown which='{which}'. Use 'first', 'last', or 'all'.")

    energy, I0 = P.build_external_baseline_avg(
        user_folder=user_folder,
        files_zero=list(files),
        which=w,
        save_npz=None,
    )
    return {"energy": np.asarray(energy, dtype=float).copy(), "I0": np.asarray(I0, dtype=float).copy()}


def load_drr_avg(
    user_folder: str,
    files: Sequence[str],
    *,
    bg_mode: str,
    y_axis: str = "auto",
    external_vector: Optional[np.ndarray] = None,
    external_energy: Optional[np.ndarray] = None,
    derivative: Optional[int] = None,
    dE_window_pts: int = 20,
    dE_polyorder: int = 2,
    dE_oversample: float = 1.0,
    dE_interp_kind: str = "cubic",
    dE_origin_like: bool = False,
    dE_pad_flat_edges: bool = True,
) -> DataCube:
    effective_y_axis = P.resolve_shared_y_axis_request(files, y_axis)
    res = P.process_ref_avg(
        user_folder=user_folder,
        files=list(files),
        bg_mode=bg_mode,
        y_axis=effective_y_axis,
        external_vector=external_vector,
        external_energy=external_energy,
        use_global_background=False,
        plot_interactive=False,
        save_png=False,
        save_dat_file=False,
        move_original=False,
        derivative=derivative,
        dE_window_pts=dE_window_pts,
        dE_polyorder=dE_polyorder,
        dE_oversample=dE_oversample,
        dE_interp_kind=dE_interp_kind,
        dE_origin_like=dE_origin_like,
        dE_pad_flat_edges=dE_pad_flat_edges,
        center_zero=True,
    )

    _validate_cube_arrays(res["energy"], res["gate_axis"], res["Z_out"], context="DRR load")

    cbar = "DR/R" if derivative is None else ("d(DR/R)/dE" if derivative == 1 else "d2(DR/R)/dE2")
    return DataCube(
        energy=np.asarray(res["energy"], dtype=float).copy(),
        gate=np.asarray(res["gate_axis"], dtype=float).copy(),
        Z=np.asarray(res["Z_out"], dtype=float).copy(),
        gate_label=res.get("gate_label", "Gate (V)"),
        title=res.get("title", "DR/R"),
        cbar_label=cbar,
    )
