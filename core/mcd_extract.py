"""Catalog, filter, and export previously processed MCD(B) analyses.

The saved MCD(B) CSV intentionally stores increasing and decreasing field
branches in separate column blocks.  This module preserves that distinction
when results are collected across experiments; repeated field values from the
return sweep are never merged with the outward sweep.
"""

from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from contextlib import contextmanager
from functools import lru_cache
import hashlib
import json
import csv
from pathlib import Path
import re
import sqlite3
from typing import Iterable, Literal, Sequence

class _LazyNumpy:
    def __getattr__(self, name: str):
        import numpy as _np
        return getattr(_np, name)


np = _LazyNumpy()


def _pandas():
    """Import pandas only when trace loading or workbook export needs it."""
    import pandas as pd
    return pd


class _LazyPandas:
    def __getattr__(self, name: str):
        return getattr(_pandas(), name)


# Compatibility handle for callers/tests that patch core.mcd_extract.pd.read_csv;
# accessing an attribute is still lazy.
pd = _LazyPandas()


McdBranch = Literal["B increasing", "B decreasing"]
BRANCHES: tuple[McdBranch, ...] = ("B increasing", "B decreasing")


@dataclass(frozen=True)
class ProcessedMcdRecord:
    record_id: str
    settings_path: Path
    trace_path: Path
    source_file: str
    package: str
    created_utc: str
    center_ev: float
    width_mev: float
    primary_metric: str
    fit_window_t: float | None
    acquisition_conditions: dict[str, tuple[float, float]]
    condition_sources: dict[str, str]
    increasing_slope_per_t: float | None
    decreasing_slope_per_t: float | None
    temperature_setpoint_k: float | None = None
    temperature_measured_k: float | None = None

    def condition_value(self, label: str) -> float | None:
        bounds = self.acquisition_conditions.get(label)
        if bounds is None:
            return None
        values = np.asarray(bounds, float)
        finite = values[np.isfinite(values)]
        return float(np.mean(finite)) if finite.size else None

    def slope(self, branch: McdBranch) -> float | None:
        return (
            self.increasing_slope_per_t
            if branch == "B increasing"
            else self.decreasing_slope_per_t
        )


@dataclass(frozen=True)
class McdExtractFilters:
    doping_v: float | None = None
    doping_tolerance_v: float = 0.01
    efield_v: float | None = None
    efield_tolerance_v: float = 0.01
    temperature_k: float | None = None
    temperature_tolerance_k: float = 0.1
    vtg_v: float | None = None
    vbg_v: float | None = None
    vbias_v: float | None = None
    gate_tolerance_v: float = 0.01
    energy_min_ev: float | None = None
    energy_max_ev: float | None = None
    width_mev: float | None = None
    width_tolerance_mev: float = 1e-3


@dataclass(frozen=True)
class McdSeries:
    """An automatically detected comparison series with per-record energies."""

    series_id: str
    variable: str
    label: str
    records: tuple[ProcessedMcdRecord, ...]
    fixed_conditions: dict[str, float | None]


_CATALOG_SCHEMA_VERSION = 4
_CATALOG_NAME = ".mcd_extract_catalog.json"
_SQLITE_CATALOG_SCHEMA_VERSION = 4
_SQLITE_CATALOG_NAME = ".mcd_catalog.sqlite3"
_SERIES_VARIABLES = ("E-field", "Temperature", "Doping", "Vtg", "Vbg", "Vbias", "Energy")
MCD_DEFAULT_TEMPERATURE_K = 1.67


def _finite_or_none(value: object) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if np.isfinite(number) else None


def _condition_mapping(value: object) -> dict[str, tuple[float, float]]:
    if not isinstance(value, dict):
        return {}
    result: dict[str, tuple[float, float]] = {}
    for label, bounds in value.items():
        if not isinstance(bounds, (list, tuple)) or len(bounds) < 2:
            continue
        low, high = _finite_or_none(bounds[0]), _finite_or_none(bounds[1])
        if low is not None and high is not None:
            result[str(label)] = (low, high)
    return result


_TEMPERATURE_KEYS = {
    "sample_tmid_k", "sample_temperature_k", "temperature_k", "temp_k", "t_k",
}


def _temperature_from_filename(source_file: str) -> float | None:
    stem = Path(source_file).stem
    match = re.search(
        r"(?:^|[_-])([+-]?\d+(?:[p.]\d+)?)\s*K(?:[_-]|$)",
        stem,
        re.IGNORECASE,
    )
    if not match:
        return None
    try:
        return float(match.group(1).replace("p", ".").replace("P", "."))
    except ValueError:
        return None


def _temperature_from_object(value: object) -> float | None:
    if isinstance(value, dict):
        for key, item in value.items():
            normalized = re.sub(r"[^a-z0-9]+", "_", str(key).casefold()).strip("_")
            if normalized in _TEMPERATURE_KEYS:
                if isinstance(item, (list, tuple)):
                    finite = [number for part in item if (number := _finite_or_none(part)) is not None]
                    if finite:
                        return float(np.mean(finite))
                number = _finite_or_none(item)
                if number is not None:
                    return number
        for item in value.values():
            found = _temperature_from_object(item)
            if found is not None:
                return found
    elif isinstance(value, list):
        for item in value:
            found = _temperature_from_object(item)
            if found is not None:
                return found
    return None


def _temperature_from_measurement_files(
    search_root: Path, source_file: str, payload: dict
) -> tuple[float | None, str]:
    """Resolve missing temperature without treating filenames as authoritative."""
    direct = _temperature_from_object(payload)
    if direct is not None:
        return direct, "MCD settings JSON"
    experiment_root = search_root.parents[1] if search_root.name.casefold() == "mcd" else search_root
    source_name = Path(source_file).name
    source_stem = Path(source_name).stem
    json_candidates: list[Path] = []
    for candidate_name in (
        f"{source_stem}.metadata.json", f"{source_name}.metadata.json", f"{source_stem}.json"
    ):
        json_candidates.extend(experiment_root.rglob(candidate_name))
    try:
        json_candidates.extend(experiment_root.rglob(f"{source_stem}*.json"))
    except OSError:
        pass
    seen: set[Path] = set()
    for candidate in json_candidates:
        try:
            resolved = candidate.resolve()
        except OSError:
            continue
        if resolved in seen or "_mcd_settings" in candidate.name.casefold():
            continue
        seen.add(resolved)
        if len(seen) > 40:
            break
        try:
            metadata = json.loads(candidate.read_text(encoding="utf-8"))
        except (OSError, UnicodeError, json.JSONDecodeError):
            continue
        temperature = _temperature_from_object(metadata)
        if temperature is not None:
            return temperature, f"measurement JSON: {candidate.name}"
    try:
        raw_candidates = list(experiment_root.rglob(source_name))
    except OSError:
        raw_candidates = []
    for candidate in raw_candidates[:20]:
        if "processed data" in {part.casefold() for part in candidate.parts}:
            continue
        try:
            header = pd.read_csv(candidate, nrows=0)
            lookup = {str(column).strip().casefold(): column for column in header.columns}
            column = next(
                (lookup[key] for key in ("sample_tmid_k", "temperature_k", "temp_k") if key in lookup),
                None,
            )
            if column is None:
                continue
            values = pd.to_numeric(pd.read_csv(candidate, usecols=[column])[column], errors="coerce")
            finite = values.to_numpy(float)
            finite = finite[np.isfinite(finite)]
            if finite.size:
                return float(np.mean(finite)), f"measurement CSV: {candidate.name}"
        except (OSError, ValueError, KeyError):
            continue
    filename_temperature = _temperature_from_filename(source_file)
    if filename_temperature is not None:
        return filename_temperature, "source filename"
    return None, "unresolved"


def _processed_mcd_root(root: str | Path) -> Path:
    candidate = Path(root).expanduser()
    for experiment_root in (candidate, candidate.parent):
        standard = experiment_root / "Processed Data" / "MCD"
        if standard.is_dir():
            return standard
    if candidate.name.casefold() == "processed data" and (candidate / "MCD").is_dir():
        return candidate / "MCD"
    return candidate


def _trace_path_from_payload(settings_path: Path, payload: dict) -> Path | None:
    outputs = payload.get("outputs", [])
    if isinstance(outputs, list):
        for output in outputs:
            name = Path(str(output)).name
            if name.casefold().endswith(".csv") and "_mcd_vs_b_" in name.casefold():
                candidate = settings_path.parent / name
                if candidate.is_file():
                    return candidate
    tag = settings_path.stem.split("_MCD_settings_", 1)
    if len(tag) == 2:
        candidates = sorted(settings_path.parent.glob(f"*_MCD_vs_B_{tag[1]}.csv"))
        if candidates:
            return candidates[0]
    return None


def _path_signature(path: Path) -> tuple[int, int] | None:
    try:
        stat = path.stat()
    except OSError:
        return None
    return int(stat.st_mtime_ns), int(stat.st_size)


@lru_cache(maxsize=2048)
def _has_complete_mcd_branches(path_text: str, modified_ns: int, size: int) -> bool:
    """Quickly reject legacy traces that do not contain both B-sweep branches."""
    del modified_ns, size
    try:
        with Path(path_text).open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            header = reader.fieldnames or []
            required = {
                "B_increasing_T", "B_decreasing_T",
                "corrected_signed_mean_increasing", "corrected_signed_mean_decreasing",
            }
            if not required.issubset({str(column).strip() for column in header}):
                return False
            # Some older exports retained the full schema but wrote no branch
            # samples. Require at least one finite point in each branch.
            found = {"inc": False, "dec": False}
            for row in reader:
                for key, columns in {
                    "inc": ("B_increasing_T", "corrected_signed_mean_increasing"),
                    "dec": ("B_decreasing_T", "corrected_signed_mean_decreasing"),
                }.items():
                    try:
                        found[key] |= all(np.isfinite(float(row.get(column, ""))) for column in columns)
                    except (TypeError, ValueError):
                        pass
                if all(found.values()):
                    return True
            return all(found.values())
    except (OSError, UnicodeError, StopIteration):
        return False


def mcd_record_has_complete_branches(record: ProcessedMcdRecord) -> bool:
    signature = _path_signature(record.trace_path)
    return signature is not None and _has_complete_mcd_branches(
        str(record.trace_path), *signature
    )


def _record_to_catalog(record: ProcessedMcdRecord, signature: tuple[int, int]) -> dict[str, object]:
    return {
        "signature": list(signature),
        "settings_path": str(record.settings_path),
        "trace_path": str(record.trace_path),
        "source_file": record.source_file,
        "package": record.package,
        "created_utc": record.created_utc,
        "center_ev": record.center_ev,
        "width_mev": record.width_mev,
        "primary_metric": record.primary_metric,
        "fit_window_t": record.fit_window_t,
        "acquisition_conditions": {
            key: list(value) for key, value in record.acquisition_conditions.items()
        },
        "condition_sources": record.condition_sources,
        "increasing_slope_per_t": record.increasing_slope_per_t,
        "decreasing_slope_per_t": record.decreasing_slope_per_t,
        "temperature_setpoint_k": record.temperature_setpoint_k,
        "temperature_measured_k": record.temperature_measured_k,
    }


def _record_from_catalog(
    value: object, *, validate_trace: bool = True
) -> ProcessedMcdRecord | None:
    if not isinstance(value, dict):
        return None
    try:
        settings_path = Path(str(value["settings_path"])).resolve()
        trace_path = Path(str(value["trace_path"])).resolve()
        center = float(value["center_ev"])
        width = float(value["width_mev"])
    except (KeyError, TypeError, ValueError, OSError):
        return None
    if (
        (validate_trace and not trace_path.is_file())
        or not np.isfinite(center)
        or not np.isfinite(width)
    ):
        return None
    sources = value.get("condition_sources", {})
    return ProcessedMcdRecord(
        record_id=str(settings_path).casefold(),
        settings_path=settings_path,
        trace_path=trace_path,
        source_file=str(value.get("source_file", settings_path.parent.name)),
        package=str(value.get("package", settings_path.parent.name)),
        created_utc=str(value.get("created_utc", "")),
        center_ev=center,
        width_mev=width,
        primary_metric=str(value.get("primary_metric", "mean")),
        fit_window_t=_finite_or_none(value.get("fit_window_t")),
        acquisition_conditions=_condition_mapping(value.get("acquisition_conditions")),
        condition_sources={
            str(key): str(source)
            for key, source in (sources.items() if isinstance(sources, dict) else [])
        },
        increasing_slope_per_t=_finite_or_none(value.get("increasing_slope_per_t")),
        decreasing_slope_per_t=_finite_or_none(value.get("decreasing_slope_per_t")),
        temperature_setpoint_k=_finite_or_none(value.get("temperature_setpoint_k")),
        temperature_measured_k=_finite_or_none(value.get("temperature_measured_k")),
    )


def _read_catalog(path: Path) -> dict[str, dict[str, object]]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError):
        return {}
    if not isinstance(payload, dict) or payload.get("schema_version") != _CATALOG_SCHEMA_VERSION:
        return {}
    entries = payload.get("entries")
    return entries if isinstance(entries, dict) else {}


def mcd_catalog_database_path(root: str | Path) -> Path:
    """Return the SQLite catalog path for an experiment or processed-MCD root."""
    return _processed_mcd_root(root) / _SQLITE_CATALOG_NAME


def _catalog_connection(database_path: Path) -> sqlite3.Connection:
    connection = sqlite3.connect(str(database_path), timeout=10.0)
    connection.execute("PRAGMA busy_timeout = 10000")
    try:
        connection.execute("PRAGMA journal_mode = WAL")
    except sqlite3.DatabaseError:
        connection.execute("PRAGMA journal_mode = DELETE")
    connection.execute("PRAGMA synchronous = NORMAL")
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS catalog_meta (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
        """
    )
    existing_schema = connection.execute(
        "SELECT value FROM catalog_meta WHERE key='schema_version'"
    ).fetchone()
    schema_changed = (
        existing_schema is not None
        and existing_schema[0] != str(_SQLITE_CATALOG_SCHEMA_VERSION)
    )
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS processed_mcd (
            settings_path TEXT PRIMARY KEY COLLATE NOCASE,
            settings_mtime_ns INTEGER NOT NULL,
            settings_size INTEGER NOT NULL,
            trace_path TEXT NOT NULL,
            source_file TEXT NOT NULL,
            created_utc TEXT NOT NULL,
            record_json TEXT NOT NULL,
            indexed_utc TEXT NOT NULL
        )
        """
    )
    if schema_changed:
        connection.execute("DELETE FROM processed_mcd")
    connection.execute(
        "CREATE INDEX IF NOT EXISTS idx_processed_mcd_source ON processed_mcd(source_file)"
    )
    connection.execute(
        "INSERT OR REPLACE INTO catalog_meta(key, value) VALUES('schema_version', ?)",
        (str(_SQLITE_CATALOG_SCHEMA_VERSION),),
    )
    if schema_changed:
        connection.execute(
            "INSERT OR REPLACE INTO catalog_meta(key, value) VALUES('catalog_complete', '0')"
        )
    return connection


@contextmanager
def _catalog_session(database_path: Path):
    connection = _catalog_connection(database_path)
    try:
        yield connection
        connection.commit()
    finally:
        connection.close()


def _upsert_catalog_record(
    connection: sqlite3.Connection, record: ProcessedMcdRecord
) -> None:
    signature = _path_signature(record.settings_path) or (0, 0)
    serialized = _record_to_catalog(record, signature)
    connection.execute(
        """
        INSERT INTO processed_mcd (
            settings_path, settings_mtime_ns, settings_size, trace_path,
            source_file, created_utc, record_json, indexed_utc
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(settings_path) DO UPDATE SET
            settings_mtime_ns=excluded.settings_mtime_ns,
            settings_size=excluded.settings_size,
            trace_path=excluded.trace_path,
            source_file=excluded.source_file,
            created_utc=excluded.created_utc,
            record_json=excluded.record_json,
            indexed_utc=excluded.indexed_utc
        """,
        (
            str(record.settings_path), signature[0], signature[1],
            str(record.trace_path), record.source_file, record.created_utc,
            json.dumps(serialized, sort_keys=True),
            datetime.now(timezone.utc).isoformat(),
        ),
    )


def _replace_sqlite_catalog(
    search_root: Path, records: Sequence[ProcessedMcdRecord]
) -> None:
    database_path = search_root / _SQLITE_CATALOG_NAME
    try:
        with _catalog_session(database_path) as connection:
            connection.execute("DELETE FROM processed_mcd")
            for record in records:
                _upsert_catalog_record(connection, record)
            connection.execute(
                "INSERT OR REPLACE INTO catalog_meta(key, value) VALUES('catalog_complete', '1')"
            )
            connection.execute(
                "INSERT OR REPLACE INTO catalog_meta(key, value) VALUES('rebuilt_utc', ?)",
                (datetime.now(timezone.utc).isoformat(),),
            )
    except (OSError, sqlite3.DatabaseError):
        return


def _load_sqlite_catalog(search_root: Path) -> list[ProcessedMcdRecord] | None:
    database_path = search_root / _SQLITE_CATALOG_NAME
    if not database_path.is_file():
        return None
    probe: sqlite3.Connection | None = None
    try:
        probe = sqlite3.connect(str(database_path), timeout=2.0)
        schema_row = probe.execute(
            "SELECT value FROM catalog_meta WHERE key='schema_version'"
        ).fetchone()
        if schema_row is None or schema_row[0] != str(_SQLITE_CATALOG_SCHEMA_VERSION):
            return None
    except (OSError, sqlite3.DatabaseError):
        return None
    finally:
        if probe is not None:
            probe.close()
    try:
        with _catalog_session(database_path) as connection:
            complete_row = connection.execute(
                "SELECT value FROM catalog_meta WHERE key='catalog_complete'"
            ).fetchone()
            if complete_row is None or complete_row[0] != "1":
                return None
            rows = connection.execute(
                "SELECT record_json FROM processed_mcd ORDER BY settings_path COLLATE NOCASE"
            ).fetchall()
    except (OSError, sqlite3.DatabaseError):
        return None
    records: list[ProcessedMcdRecord] = []
    for (record_text,) in rows:
        try:
            value = json.loads(record_text)
        except (TypeError, json.JSONDecodeError):
            continue
        record = _record_from_catalog(value, validate_trace=False)
        if record is not None:
            records.append(record)
    return records


def _slope_from_csv(path: Path, column: str) -> float | None:
    try:
        values = pd.read_csv(path, usecols=[column])[column].to_numpy(float)
    except (OSError, ValueError, KeyError):
        return None
    finite = values[np.isfinite(values)]
    return float(finite[0]) if finite.size else None


def _record_from_settings_payload(
    search_root: Path, settings_path: Path, payload: dict[str, object]
) -> ProcessedMcdRecord | None:
    if str(payload.get("workflow", "")).casefold() != "mcd":
        return None
    mcd_b = payload.get("mcd_b")
    if not isinstance(mcd_b, dict):
        return None
    center = _finite_or_none(mcd_b.get("center_ev"))
    width = _finite_or_none(mcd_b.get("width_mev"))
    trace_path = _trace_path_from_payload(settings_path, payload)
    if center is None or width is None or trace_path is None:
        return None
    increasing = _finite_or_none(mcd_b.get("low_field_mcd_slope_increasing_per_T"))
    decreasing = _finite_or_none(mcd_b.get("low_field_mcd_slope_decreasing_per_T"))
    if increasing is None:
        increasing = _slope_from_csv(trace_path, "low_field_mcd_slope_increasing_per_T")
    if decreasing is None:
        decreasing = _slope_from_csv(trace_path, "low_field_mcd_slope_decreasing_per_T")
    fit_window = _finite_or_none(mcd_b.get("fit_window_t"))
    source_file = str(payload.get("source_file", settings_path.parent.name))
    conditions = _condition_mapping(payload.get("acquisition_conditions"))
    condition_sources = {label: "MCD settings JSON" for label in conditions}
    measured_temperature = None
    if "T" in conditions:
        bounds = conditions["T"]
        finite = [value for value in bounds if np.isfinite(value)]
        measured_temperature = float(np.mean(finite)) if finite else None
    filename_temperature = _temperature_from_filename(source_file)
    if filename_temperature is not None:
        # The filename describes the requested cryostat setpoint.  Metadata
        # remains useful as a measured diagnostic, but must not split a
        # nominal 2.5 K sweep into 2.53/2.67/2.73 K groups.
        if measured_temperature is None:
            discovered, discovered_source = _temperature_from_measurement_files(
                search_root, source_file, payload
            )
            if discovered is not None and discovered_source != "source filename":
                measured_temperature = discovered
        conditions["T"] = (filename_temperature, filename_temperature)
        condition_sources["T"] = "source filename setpoint"
    elif "T" not in conditions:
        temperature, temperature_source = _temperature_from_measurement_files(
            search_root, source_file, payload
        )
        if temperature is not None:
            conditions["T"] = (temperature, temperature)
            condition_sources["T"] = temperature_source
            measured_temperature = temperature
        else:
            conditions["T"] = (
                MCD_DEFAULT_TEMPERATURE_K,
                MCD_DEFAULT_TEMPERATURE_K,
            )
            condition_sources["T"] = "assumed default: missing temperature metadata"
    resolved_settings = settings_path.resolve()
    return ProcessedMcdRecord(
        record_id=str(resolved_settings).casefold(),
        settings_path=resolved_settings,
        trace_path=trace_path.resolve(),
        source_file=source_file,
        package=str(payload.get("package", settings_path.parent.name)),
        created_utc=str(payload.get("created_utc", "")),
        center_ev=center,
        width_mev=width,
        primary_metric=str(mcd_b.get("primary_metric", "mean")),
        fit_window_t=fit_window if bool(mcd_b.get("fit_near_zero", False)) else None,
        acquisition_conditions=conditions,
        condition_sources=condition_sources,
        increasing_slope_per_t=increasing,
        decreasing_slope_per_t=decreasing,
        temperature_setpoint_k=filename_temperature,
        temperature_measured_k=measured_temperature,
    )


def _processed_root_for_settings(settings_path: Path) -> Path | None:
    for parent in settings_path.resolve().parents:
        if (
            parent.name.casefold() == "mcd"
            and parent.parent.name.casefold() == "processed data"
        ):
            return parent
    return None


def index_processed_mcd_settings(
    settings_path: str | Path, *, payload: dict[str, object] | None = None
) -> bool:
    """Insert or update one saved MCD analysis in the SQLite catalog."""
    path = Path(settings_path).resolve()
    search_root = _processed_root_for_settings(path)
    if search_root is None or not path.is_file():
        return False
    if payload is None:
        try:
            loaded = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, UnicodeError, json.JSONDecodeError):
            return False
        if not isinstance(loaded, dict):
            return False
        payload = loaded
    record = _record_from_settings_payload(search_root, path, payload)
    if record is None or not mcd_record_has_complete_branches(record):
        return False
    try:
        with _catalog_session(search_root / _SQLITE_CATALOG_NAME) as connection:
            _upsert_catalog_record(connection, record)
    except (OSError, sqlite3.DatabaseError):
        return False
    return True


def discover_processed_mcd(
    root: str | Path, *, rebuild_catalog: bool = False
) -> list[ProcessedMcdRecord]:
    """Find processed results using an incremental metadata catalog."""
    search_root = _processed_mcd_root(root)
    if not search_root.is_dir():
        return []
    if not rebuild_catalog:
        indexed_records = _load_sqlite_catalog(search_root)
        if indexed_records is not None:
            return sorted(
                indexed_records,
                key=lambda item: (
                    item.condition_value("Doping") is None,
                    item.condition_value("Doping") or 0.0,
                    item.condition_value("E-field") is None,
                    item.condition_value("E-field") or 0.0,
                    item.center_ev,
                    item.width_mev,
                    item.created_utc,
                ),
            )
    catalog_path = search_root / _CATALOG_NAME
    cached = {} if rebuild_catalog else _read_catalog(catalog_path)
    records: list[ProcessedMcdRecord] = []
    for settings_path in sorted(search_root.rglob("*_MCD_settings*.json")):
        resolved_settings = settings_path.resolve()
        cache_key = str(resolved_settings).casefold()
        signature = _path_signature(settings_path)
        cached_entry = cached.get(cache_key)
        cached_signature = cached_entry.get("signature") if isinstance(cached_entry, dict) else None
        if signature is not None and cached_signature == list(signature):
            record = _record_from_catalog(cached_entry)
            if record is not None:
                records.append(record)
                continue
        try:
            payload = json.loads(settings_path.read_text(encoding="utf-8"))
        except (OSError, UnicodeError, json.JSONDecodeError):
            continue
        if not isinstance(payload, dict):
            continue
        record = _record_from_settings_payload(search_root, settings_path, payload)
        if record is None:
            continue
        records.append(record)
    records = [record for record in records if mcd_record_has_complete_branches(record)]
    _replace_sqlite_catalog(search_root, records)
    return sorted(
        records,
        key=lambda item: (
            item.condition_value("Doping") is None,
            item.condition_value("Doping") or 0.0,
            item.condition_value("E-field") is None,
            item.condition_value("E-field") or 0.0,
            item.center_ev,
            item.width_mev,
            item.created_utc,
        ),
    )


def filter_processed_mcd(
    records: Iterable[ProcessedMcdRecord], filters: McdExtractFilters
) -> list[ProcessedMcdRecord]:
    """Apply tolerant numeric filters to a processed-result catalog."""
    selected: list[ProcessedMcdRecord] = []
    for record in records:
        doping = record.condition_value("Doping")
        efield = record.condition_value("E-field")
        if filters.doping_v is not None and (
            doping is None or abs(doping - filters.doping_v) > abs(filters.doping_tolerance_v)
        ):
            continue
        if filters.efield_v is not None and (
            efield is None or abs(efield - filters.efield_v) > abs(filters.efield_tolerance_v)
        ):
            continue
        additional_conditions = (
            ("T", filters.temperature_k, filters.temperature_tolerance_k),
            ("Vtg", filters.vtg_v, filters.gate_tolerance_v),
            ("Vbg", filters.vbg_v, filters.gate_tolerance_v),
            ("Vbias", filters.vbias_v, filters.gate_tolerance_v),
        )
        condition_failed = False
        for label, target, tolerance in additional_conditions:
            if target is None:
                continue
            actual = record.condition_value(label)
            if actual is None or abs(actual - target) > abs(tolerance):
                condition_failed = True
                break
        if condition_failed:
            continue
        if filters.energy_min_ev is not None and record.center_ev < filters.energy_min_ev:
            continue
        if filters.energy_max_ev is not None and record.center_ev > filters.energy_max_ev:
            continue
        if filters.width_mev is not None and (
            abs(record.width_mev - filters.width_mev) > abs(filters.width_tolerance_mev)
        ):
            continue
        selected.append(record)
    return selected


def energy_cluster_centers(
    records: Sequence[ProcessedMcdRecord], tolerance_mev: float
) -> dict[str, float]:
    """Assign nearby integration centers to stable, non-chaining energy groups."""
    tolerance_ev = max(0.0, float(tolerance_mev)) * 1e-3
    groups: list[list[ProcessedMcdRecord]] = []
    for record in sorted(records, key=lambda item: item.center_ev):
        placed = False
        for group in groups:
            center = float(np.mean([item.center_ev for item in group]))
            if abs(record.center_ev - center) <= tolerance_ev:
                group.append(record)
                placed = True
                break
        if not placed:
            groups.append([record])
    result: dict[str, float] = {}
    for group in groups:
        center = float(np.mean([item.center_ev for item in group]))
        for record in group:
            result[record.record_id] = center
    return result


def _created_sort_value(record: ProcessedMcdRecord) -> tuple[float, int]:
    text = record.created_utc.strip().replace("Z", "+00:00")
    try:
        timestamp = datetime.fromisoformat(text).timestamp()
    except (TypeError, ValueError, OSError):
        timestamp = 0.0
    signature = _path_signature(record.settings_path)
    return timestamp, signature[0] if signature else 0


def newest_mcd_versions(
    records: Sequence[ProcessedMcdRecord],
) -> tuple[list[ProcessedMcdRecord], list[ProcessedMcdRecord]]:
    """Keep the newest reprocessing of each source/energy/width/metric identity."""
    newest: dict[tuple[str, float, float, str], ProcessedMcdRecord] = {}
    older: list[ProcessedMcdRecord] = []
    for record in records:
        key = (
            Path(record.source_file).as_posix().casefold(),
            round(float(record.center_ev), 9),
            round(float(record.width_mev), 6),
            record.primary_metric.casefold(),
        )
        previous = newest.get(key)
        if previous is None or _created_sort_value(record) > _created_sort_value(previous):
            if previous is not None:
                older.append(previous)
            newest[key] = record
        else:
            older.append(record)
    kept = sorted(
        newest.values(),
        key=lambda item: (item.source_file.casefold(), item.center_ev, item.width_mev),
    )
    return kept, older


def _series_value(record: ProcessedMcdRecord, variable: str) -> float | None:
    if variable == "Energy":
        return record.center_ev
    return record.condition_value("T" if variable == "Temperature" else variable)


def organize_mcd_series(
    records: Sequence[ProcessedMcdRecord], variable: str = "Auto", *,
    include_singletons: bool = True,
) -> list[McdSeries]:
    """Partition results by fixed conditions while preserving each saved energy."""
    if not records:
        return []
    candidates = [variable] if variable in _SERIES_VARIABLES else list(_SERIES_VARIABLES)
    condition_variables = ("Doping", "E-field", "Temperature", "Vtg", "Vbg", "Vbias")
    tolerances = {
        "Doping": 0.01, "E-field": 0.01, "Temperature": 0.1,
        "Vtg": 0.01, "Vbg": 0.01, "Vbias": 0.01,
    }

    def fixed_variables_for_axis(axis: str) -> tuple[str, ...]:
        # Doping and E-field are physical coordinates derived from Vtg/Vbg.
        # Holding the raw gates fixed while comparing either coordinate makes
        # a valid dual-gate sweep impossible.  Raw gate comparisons instead
        # hold the opposite gate fixed and do not constrain the derived axes.
        return {
            "E-field": ("Doping", "Temperature", "Vbias"),
            "Temperature": ("Doping", "E-field", "Vbias"),
            "Doping": ("E-field", "Temperature", "Vbias"),
            "Vtg": ("Vbg", "Temperature", "Vbias"),
            "Vbg": ("Vtg", "Temperature", "Vbias"),
            "Vbias": ("Doping", "E-field", "Temperature"),
            "Energy": condition_variables,
        }.get(axis, condition_variables)

    def partition(axis: str) -> list[list[ProcessedMcdRecord]]:
        fixed_variables = fixed_variables_for_axis(axis)
        groups: list[list[ProcessedMcdRecord]] = []
        for record in records:
            matched: list[ProcessedMcdRecord] | None = None
            for group in groups:
                representative = group[0]
                same = abs(record.width_mev - representative.width_mev) <= 1e-3
                for item in fixed_variables:
                    left = _series_value(record, item)
                    right = _series_value(representative, item)
                    if left is None or right is None:
                        same = same and left is None and right is None
                    else:
                        same = same and abs(left - right) <= tolerances[item]
                    if not same:
                        break
                if same:
                    matched = group
                    break
            if matched is None:
                groups.append([record])
            else:
                matched.append(record)
        return groups

    def distinct_axis_values(group: Sequence[ProcessedMcdRecord], axis: str) -> set[float]:
        return {
            float(value) for item in group
            if (value := _series_value(item, axis)) is not None and np.isfinite(value)
        }

    best_axis = candidates[0]
    best_groups = partition(best_axis)
    best_score = (-1, -1, -1)
    for priority, axis in enumerate(candidates):
        groups = partition(axis)
        multi = [
            group for group in groups
            if len(distinct_axis_values(group, axis)) >= 2
        ]
        score = (
            sum(len(group) for group in multi),
            max((len(group) for group in multi), default=0),
            -priority,
        )
        if score > best_score:
            best_axis, best_groups, best_score = axis, groups, score

    if not include_singletons:
        best_groups = [
            group for group in best_groups
            if len(distinct_axis_values(group, best_axis)) >= 2
        ]

    series: list[McdSeries] = []
    for index, group in enumerate(best_groups, start=1):
        ordered, _ = order_mcd_records(group, best_axis)
        fixed: dict[str, float | None] = {}
        for item in fixed_variables_for_axis(best_axis):
            values = [_series_value(record, item) for record in group]
            finite = [value for value in values if value is not None and np.isfinite(value)]
            fixed[item] = float(np.mean(finite)) if finite else None
        required_fixed = {
            "E-field": ("Doping", "Temperature"),
            "Temperature": ("Doping", "E-field"),
        }.get(best_axis, ())
        fixed_parts: list[str] = []
        for name, value in fixed.items():
            if value is not None:
                fixed_parts.append(f"{name}={value:.6g}")
            elif name in required_fixed:
                fixed_parts.append(f"{name}=?")
        fixed_text = ", ".join(fixed_parts)
        energy_low = min(record.center_ev for record in group)
        energy_high = max(record.center_ev for record in group)
        energy_text = (
            f"E={energy_low:.6g} eV" if np.isclose(energy_low, energy_high)
            else f"E={energy_low:.6g}–{energy_high:.6g} eV"
        )
        axis_values = [
            value for record in group
            if (value := _series_value(record, best_axis)) is not None and np.isfinite(value)
        ]
        axis_units = ORDER_VARIABLES.get(best_axis, (best_axis, "", None))[1]
        if axis_values:
            axis_low, axis_high = min(axis_values), max(axis_values)
            axis_text = (
                f"{best_axis}={axis_low:.6g}{axis_units}"
                if np.isclose(axis_low, axis_high)
                else f"{best_axis}={axis_low:.6g}→{axis_high:.6g}{axis_units}"
            )
        else:
            axis_text = f"{best_axis}=unknown"
        label = (
            f"{best_axis} series ({len(group)}) · {axis_text} · "
            f"{fixed_text or 'mixed conditions'} · {energy_text}"
        )
        digest = hashlib.sha256(
            "|".join(record.record_id for record in ordered).encode("utf-8")
        ).hexdigest()[:10]
        series.append(
            McdSeries(f"S{index:02d}_{digest}", best_axis, label, tuple(ordered), fixed)
        )
    return sorted(series, key=lambda item: (-len(item.records), item.label.casefold()))


@lru_cache(maxsize=256)
def _cached_trace_table(path_text: str, modified_ns: int, size: int) -> pd.DataFrame:
    del modified_ns, size
    return pd.read_csv(path_text)


def clear_mcd_trace_cache() -> None:
    _cached_trace_table.cache_clear()


def load_branch_traces(
    record: ProcessedMcdRecord,
    branches: Sequence[McdBranch] = BRANCHES,
) -> pd.DataFrame:
    """Return tidy trace rows while keeping the two acquired branches separate."""
    signature = _path_signature(record.trace_path)
    if signature is None:
        raise OSError(f"MCD trace file is unavailable: {record.trace_path}")
    table = _cached_trace_table(str(record.trace_path), *signature)
    blocks: list[pd.DataFrame] = []
    for branch in branches:
        suffix = "increasing" if branch == "B increasing" else "decreasing"
        required = {
            "B_T": f"B_{suffix}_T",
            "corrected_signed_mean": f"corrected_signed_mean_{suffix}",
            "corrected_field_signed_absolute_mean": f"corrected_field_signed_absolute_mean_{suffix}",
            "corrected_integral": f"corrected_integral_{suffix}",
        }
        missing = [column for column in required.values() if column not in table.columns]
        if missing:
            raise ValueError(f"{record.trace_path.name} is missing columns: {', '.join(missing)}")
        block = table[list(required.values())].rename(columns={value: key for key, value in required.items()})
        block = block[np.isfinite(pd.to_numeric(block["B_T"], errors="coerce"))].copy()
        block.insert(0, "branch", branch)
        block.insert(0, "width_mev", record.width_mev)
        block.insert(0, "energy_ev", record.center_ev)
        block.insert(0, "efield_V", record.condition_value("E-field"))
        block.insert(0, "doping_V", record.condition_value("Doping"))
        block.insert(0, "source_file", record.source_file)
        block.insert(0, "record_id", record.record_id)
        blocks.append(block)
    return pd.concat(blocks, ignore_index=True) if blocks else pd.DataFrame()


def slope_summary(
    records: Sequence[ProcessedMcdRecord],
    branches: Sequence[McdBranch] = BRANCHES,
) -> pd.DataFrame:
    """Create one slope row per processed result and field-sweep branch."""
    rows: list[dict[str, object]] = []
    for record in records:
        traces = load_branch_traces(record, branches)
        for branch in branches:
            branch_data = traces[traces["branch"] == branch]
            b_values = pd.to_numeric(branch_data.get("B_T"), errors="coerce").to_numpy(float)
            y_values = pd.to_numeric(
                branch_data.get("corrected_signed_mean"), errors="coerce"
            ).to_numpy(float)
            mask = np.isfinite(b_values) & np.isfinite(y_values)
            if record.fit_window_t is not None:
                mask &= np.abs(b_values) <= abs(record.fit_window_t)
            intercept = r_squared = float("nan")
            calculated_slope = float("nan")
            count = int(np.count_nonzero(mask))
            if count >= 2 and np.ptp(b_values[mask]) > 0:
                calculated_slope, intercept = np.polyfit(b_values[mask], y_values[mask], 1)
                fitted = calculated_slope * b_values[mask] + intercept
                residual = float(np.sum((y_values[mask] - fitted) ** 2))
                total = float(np.sum((y_values[mask] - np.mean(y_values[mask])) ** 2))
                r_squared = 1.0 - residual / total if total > 0 else (1.0 if residual == 0 else np.nan)
            stored_slope = record.slope(branch)
            rows.append({
                "record_id": record.record_id,
                "source_file": record.source_file,
                "package": record.package,
                "doping_V": record.condition_value("Doping"),
                "efield_V": record.condition_value("E-field"),
                "temperature_K": record.condition_value("T"),
                "energy_ev": record.center_ev,
                "width_mev": record.width_mev,
                "branch": branch,
                "slope_per_T": stored_slope if stored_slope is not None else calculated_slope,
                "intercept": intercept,
                "fit_half_range_T": record.fit_window_t,
                "fit_point_count": count,
                "fit_r_squared": r_squared,
                "created_utc": record.created_utc,
                "settings_file": record.settings_path.name,
            })
    return pd.DataFrame(rows)


ORDER_VARIABLES: dict[str, tuple[str, str, str | None]] = {
    "E-field": ("E-field", "V", "E-field"),
    "Temperature": ("Temperature", "K", "T"),
    "Doping": ("Doping", "V", "Doping"),
    "Energy": ("Energy", "eV", None),
    "Vtg": ("Vtg", "V", "Vtg"),
    "Vbg": ("Vbg", "V", "Vbg"),
    "Vbias": ("Vbias", "V", "Vbias"),
    "Width": ("Width", "meV", None),
    "Source": ("Source", "", None),
}

PALETTES: tuple[str, ...] = (
    "viridis", "plasma", "inferno", "magma", "cividis", "turbo", "cubehelix",
    "Blues", "BuGn", "BuPu", "GnBu", "Greens", "Greys", "Oranges", "OrRd",
    "PuBu", "PuBuGn", "PuRd", "Purples", "RdPu", "Reds", "YlGn", "YlGnBu",
    "YlOrBr", "YlOrRd", "coolwarm", "Spectral", "RdBu", "RdYlBu", "RdYlGn",
    "PiYG", "PRGn", "BrBG", "PuOr", "seismic", "twilight", "twilight_shifted",
)


def record_order_value(record: ProcessedMcdRecord, variable: str) -> float | str | None:
    if variable == "Energy":
        return record.center_ev
    if variable == "Width":
        return record.width_mev
    if variable == "Source":
        return Path(record.source_file).stem.casefold()
    descriptor = ORDER_VARIABLES.get(variable)
    return record.condition_value(descriptor[2]) if descriptor and descriptor[2] else None


def _varies(records: Sequence[ProcessedMcdRecord], variable: str) -> bool:
    values = [record_order_value(record, variable) for record in records]
    finite = [float(value) for value in values if isinstance(value, (int, float)) and np.isfinite(value)]
    return len(finite) >= 2 and not np.allclose(finite, finite[0], rtol=0, atol=1e-9)


def resolve_order_variable(records: Sequence[ProcessedMcdRecord], requested: str = "Auto") -> str:
    if requested in ORDER_VARIABLES:
        return requested
    doping_fixed = not _varies(records, "Doping")
    efield_fixed = not _varies(records, "E-field")
    if doping_fixed and _varies(records, "E-field"):
        return "E-field"
    if doping_fixed and efield_fixed and _varies(records, "Temperature"):
        return "Temperature"
    for variable in ("Energy", "Doping", "Vtg", "Vbg", "Vbias", "Width"):
        if _varies(records, variable):
            return variable
    return "Source"


def order_mcd_records(
    records: Sequence[ProcessedMcdRecord], requested: str = "Auto", *, descending: bool = False
) -> tuple[list[ProcessedMcdRecord], str]:
    resolved = resolve_order_variable(records, requested)

    def key(record: ProcessedMcdRecord) -> tuple[bool, object, float, str]:
        value = record_order_value(record, resolved)
        missing = value is None or (isinstance(value, float) and not np.isfinite(value))
        sortable: object = "" if value is None else value
        return missing, sortable, record.center_ev, record.source_file.casefold()

    known = [record for record in records if key(record)[0] is False]
    unknown = [record for record in records if key(record)[0] is True]
    known.sort(key=key, reverse=bool(descending))
    unknown.sort(key=lambda record: (record.center_ev, record.source_file.casefold()))
    return known + unknown, resolved


def isolated_condition_values(values: Sequence[float]) -> set[float]:
    """Find single condition values separated far beyond the normal series spacing."""
    unique = np.asarray(sorted({float(value) for value in values if np.isfinite(value)}), float)
    if unique.size < 3:
        return set()
    gaps = np.diff(unique)
    positive = np.sort(gaps[gaps > 1e-12])
    if positive.size < 2:
        return set()
    baseline = float(np.median(positive[:-1]))
    if baseline <= 0 or float(positive[-1]) < 3.0 * baseline:
        return set()
    isolated: set[float] = set()
    for index, value in enumerate(unique):
        neighbor_gaps = []
        if index:
            neighbor_gaps.append(value - unique[index - 1])
        if index + 1 < unique.size:
            neighbor_gaps.append(unique[index + 1] - value)
        if neighbor_gaps and min(neighbor_gaps) >= 3.0 * baseline:
            isolated.add(float(value))
    return isolated


def assign_plot_colors(
    records: Sequence[ProcessedMcdRecord], palette: str, variable: str | None = None,
) -> dict[str, str]:
    from matplotlib import colormaps
    from matplotlib.colors import Normalize, TwoSlopeNorm, to_hex

    selected_palette = palette if palette in PALETTES else "viridis"
    cmap = colormaps[selected_palette]
    values = [record_order_value(record, variable) if variable else None for record in records]
    numeric = np.asarray([
        float(value) if isinstance(value, (int, float)) and np.isfinite(value) else np.nan
        for value in values
    ], float)
    finite = numeric[np.isfinite(numeric)]
    isolated = isolated_condition_values(finite) if variable and variable != "Source" else set()
    mapped = np.asarray([
        value for value in finite if not any(np.isclose(value, item, atol=1e-12) for item in isolated)
    ], float)
    if mapped.size:
        low, high = float(np.min(mapped)), float(np.max(mapped))
        if low < 0.0 < high:
            norm = TwoSlopeNorm(vmin=low, vcenter=0.0, vmax=high)
        elif high > low:
            norm = Normalize(vmin=low, vmax=high)
        else:
            padding = max(abs(low) * 0.05, 0.5)
            norm = Normalize(vmin=low - padding, vmax=high + padding)
        positions = [
            0.5 if not np.isfinite(value) else float(norm(float(value)))
            for value in numeric
        ]
    else:
        count = len(records)
        positions = [0.5] if count == 1 else np.linspace(0.08, 0.92, count)
    return {
        record.record_id: (
            "#4a4a4a" if variable and variable != "Source" and np.isfinite(numeric[index])
            and any(np.isclose(numeric[index], item, atol=1e-12) for item in isolated)
            else to_hex(cmap(float(np.clip(position, 0.0, 1.0))), keep_alpha=False)
        )
        for index, (record, position) in enumerate(zip(records, positions))
    }


def _display_condition_value(value: object, variable: str) -> str:
    """Return a compact, human-readable value for legends and headers."""
    if value is None:
        return "?"
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    decimals = {
        "Energy": 3, "Doping": 3, "E-field": 3, "Temperature": 3,
        "Vtg": 3, "Vbg": 3, "Vbias": 3, "Width": 3,
    }.get(variable, 3)
    return f"{number:.{decimals}g}"


def concise_condition_labels(
    records: Sequence[ProcessedMcdRecord], order_variable: str,
) -> tuple[list[str], str]:
    """Build short labels from measurement conditions, never source filenames."""
    variables = ("E-field", "Energy", "Doping", "Temperature", "Vtg", "Vbg", "Vbias", "Width")

    def value(record: ProcessedMcdRecord, variable: str) -> float | None:
        return record.center_ev if variable == "Energy" else record.condition_value(variable)

    varying = [
        variable for variable in variables
        if len({_display_condition_value(value(record, variable), variable) for record in records}) > 1
    ]
    if order_variable != "Source" and order_variable not in varying and len(records) > 1:
        varying.insert(0, order_variable)
    labels: list[str] = []
    for record in records:
        parts = []
        for variable in varying:
            number = value(record, variable)
            if number is not None:
                parts.append(f"{variable} = {_display_condition_value(number, variable)}")
        labels.append(" | ".join(parts) or "Selected condition")

    fixed_parts = []
    for variable in ("Doping", "E-field", "Temperature", "Width"):
        vals = {_display_condition_value(value(record, variable), variable) for record in records}
        if len(vals) == 1 and "?" not in vals:
            fixed_parts.append(f"{variable} = {next(iter(vals))}")
    return labels, " | ".join(fixed_parts)


def _curve_label(record: ProcessedMcdRecord, order_variable: str, index: int) -> str:
    value = record_order_value(record, order_variable)
    unit = ORDER_VARIABLES[order_variable][1]
    value_text = "unknown" if value is None else _display_condition_value(value, order_variable)
    abbreviations = {
        "E-field": "F", "Temperature": "T", "Doping": "D", "Energy": "E",
        "Vtg": "Vtg", "Vbg": "Vbg", "Vbias": "Vbias", "Width": "W", "Source": "S",
    }
    if order_variable == "Source":
        value_text = re.sub(r"[^A-Za-z0-9]+", "_", str(value_text)).strip("_")[:24]
    energy_text = f"{record.center_ev:.6g}"
    if order_variable == "Energy":
        return f"E{energy_text}eV"
    condition_unit = "" if order_variable == "E-field" else unit
    return f"{abbreviations[order_variable]}{value_text}{condition_unit}_E{energy_text}eV"


def origin_branch_table(
    records: Sequence[ProcessedMcdRecord], branch: McdBranch, order_variable: str
) -> pd.DataFrame:
    """Build repeated, simple Origin-friendly X/Y blocks for one sweep branch."""
    columns: dict[str, pd.Series] = {}
    used_prefixes: dict[str, int] = {}
    for index, record in enumerate(records, start=1):
        trace = load_branch_traces(record, (branch,))
        b_values = pd.to_numeric(trace["B_T"], errors="coerce").to_numpy(float)
        mcd_values = pd.to_numeric(trace["corrected_signed_mean"], errors="coerce").to_numpy(float)
        prefix = _curve_label(record, order_variable, index)
        used_prefixes[prefix] = used_prefixes.get(prefix, 0) + 1
        if used_prefixes[prefix] > 1:
            prefix = f"{prefix}_{used_prefixes[prefix]}"
        columns[f"{prefix}_B_T"] = pd.Series(b_values)
        columns[f"{prefix}_MCD"] = pd.Series(mcd_values)
    return pd.DataFrame(columns)


def _excel_value(value: object) -> object:
    if value is None or value is pd.NA:
        return None
    if isinstance(value, (np.integer, np.floating)):
        value = value.item()
    if isinstance(value, float) and not np.isfinite(value):
        return None
    return value


def _write_dataframe_sheet(workbook: Workbook, name: str, table: pd.DataFrame) -> None:
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    sheet = workbook.create_sheet(name)
    sheet.freeze_panes = "A2"
    sheet.append([str(column) for column in table.columns])
    for row in table.itertuples(index=False, name=None):
        sheet.append([_excel_value(value) for value in row])
    for cell in sheet[1]:
        cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 22
    for column_index, column_name in enumerate(table.columns, start=1):
        letter = get_column_letter(column_index)
        sheet.column_dimensions[letter].width = min(34, max(12, len(str(column_name)) + 2))
        if any(token in str(column_name) for token in ("_B_T", "_MCD", "slope", "intercept", "R2")):
            for cell in sheet[letter][1:]:
                cell.number_format = "0.000000E+00"
    if table.shape[1]:
        sheet.auto_filter.ref = f"A1:{get_column_letter(table.shape[1])}{table.shape[0] + 1}"


def _sample_token(records: Sequence[ProcessedMcdRecord]) -> str:
    token_lists = [re.split(r"[_\s-]+", Path(record.source_file).stem) for record in records]
    shared: list[str] = []
    for tokens in zip(*token_lists):
        if len({token.casefold() for token in tokens}) != 1:
            break
        token = tokens[0]
        if re.match(r"^(?:D|F|E|T|Vtg|Vbg|Vbias)[=+-]?\d", token, re.IGNORECASE):
            break
        shared.append(token)
    text = "_".join(shared).strip("_") or (Path(records[0].source_file).stem if len(records) == 1 else "MultiSample")
    return re.sub(r"[^A-Za-z0-9._-]+", "_", text).strip("._-") or "MultiSample"


def _range_token(records: Sequence[ProcessedMcdRecord], variable: str, prefix: str, unit: str) -> str:
    values = [record_order_value(record, variable) for record in records]
    finite = sorted({float(value) for value in values if isinstance(value, (int, float)) and np.isfinite(value)})
    if not finite:
        return ""
    number = lambda value: f"{value:.6g}"
    text = number(finite[0]) if len(finite) == 1 else f"{number(finite[0])}to{number(finite[-1])}"
    return f"{prefix}{text}{unit}"


def descriptive_export_base(
    records: Sequence[ProcessedMcdRecord], order_variable: str, descending: bool
) -> str:
    parts = ["MCD", _sample_token(records)]
    for variable, prefix, unit in (
        ("Doping", "D", "V"), ("E-field", "F", "V"), ("Temperature", "T", "K"),
        ("Energy", "E", "eV"), ("Width", "W", "meV"),
    ):
        token = _range_token(records, variable, prefix, unit)
        if token:
            parts.append(token)
    parts.extend([f"by{order_variable.replace('-', '')}", "desc" if descending else "asc"])
    base = re.sub(r"[^A-Za-z0-9._-]+", "_", "_".join(parts)).strip("._-")
    if len(base) > 170:
        digest = hashlib.sha256(base.encode("utf-8")).hexdigest()[:8]
        base = f"{base[:160].rstrip('._-')}_{digest}"
    return base


def _unused_export_base(out: Path, requested: str) -> str:
    suffixes = (
        "_Origin.xlsx", "_Summary.xlsx", "_MCD_increasing_vs_B.png", "_MCD_decreasing_vs_B.png",
        "_Slope_vs_Efield.png", "_Slope_vs_Temperature.png", "_settings.json",
    )
    if not any((out / f"{requested}{suffix}").exists() for suffix in suffixes):
        return requested
    for number in range(2, 10000):
        candidate = f"{requested}_{number:02d}"
        if not any((out / f"{candidate}{suffix}").exists() for suffix in suffixes):
            return candidate
    raise FileExistsError("Could not create an unused MCD extract filename.")


def _export_comparison_png(
    path: Path,
    records: Sequence[ProcessedMcdRecord],
    branches: Sequence[McdBranch],
    colors: dict[str, str],
    order_variable: str,
    palette: str,
) -> None:
    from matplotlib.backends.backend_agg import FigureCanvasAgg
    from matplotlib.cm import ScalarMappable
    from matplotlib import colormaps
    from matplotlib.colors import Normalize, TwoSlopeNorm
    from matplotlib.figure import Figure
    from matplotlib.lines import Line2D
    from matplotlib.patches import Patch

    branch_list = list(branches)
    figure = Figure(figsize=(6.0, 4.5), dpi=300, facecolor="white")
    FigureCanvasAgg(figure)
    axes = [figure.add_axes([0.15, 0.14, 0.68, 0.76])]
    if len(branch_list) > 1:
        axes = [figure.add_axes([0.10, 0.14, 0.36, 0.76]),
                figure.add_axes([0.50, 0.14, 0.36, 0.76])]

    centers = energy_cluster_centers(records, 5.0)
    energy_values = sorted({round(value, 9) for value in centers.values()})
    linestyles = ("-", "--", "-.", ":")
    markers = ("o", "s", "^", "D", "v", "P", "X")
    energy_styles = {
        value: (linestyles[index % len(linestyles)], markers[index % len(markers)])
        for index, value in enumerate(energy_values)
    }
    all_y: list[np.ndarray] = []
    for record in records:
        traces = load_branch_traces(record, branches)
        energy = round(centers.get(record.record_id, record.center_ev), 9)
        linestyle, marker = energy_styles[energy]
        for axis, branch in zip(axes, branch_list):
            block = traces[traces["branch"] == branch]
            if block.empty:
                continue
            y_values = block["corrected_signed_mean"].to_numpy(float)
            finite_y = y_values[np.isfinite(y_values)]
            if finite_y.size:
                all_y.append(finite_y)
            axis.plot(
                block["B_T"], y_values,
                linestyle=linestyle, marker=marker, markersize=2.7,
                markevery=max(1, len(block) // 24), linewidth=1.25,
                color=colors[record.record_id], markerfacecolor="white",
                markeredgecolor=colors[record.record_id],
                markeredgewidth=0.7, label="_nolegend_",
            )
    if all_y:
        combined = np.concatenate(all_y)
        low_y, high_y = float(np.min(combined)), float(np.max(combined))
        padding = max(0.004, 0.045 * max(high_y - low_y, 1e-12))
        for axis in axes:
            axis.set_ylim(low_y - padding, high_y + padding)

    for index, (axis, branch) in enumerate(zip(axes, branch_list)):
        axis.axhline(0.0, color="#555555", linewidth=0.65)
        axis.axvline(0.0, color="#777777", linewidth=0.45, alpha=0.55)
        axis.set_xlabel("B field (T)", fontsize=12)
        if index == 0:
            axis.set_ylabel("Corrected signed-mean MCD", fontsize=12)
        fixed_bits = []
        for label, key, unit in (("D", "Doping", "V"), ("T", "T", "K")):
            values = [record.condition_value(key) for record in records]
            finite = [float(value) for value in values if value is not None and np.isfinite(value)]
            if finite and np.allclose(finite, finite[0], atol=1e-6, rtol=0):
                fixed_bits.append(f"{label} = {finite[0]:g} {unit}")
        branch_name = "increasing" if branch == "B increasing" else "decreasing"
        suffix = f" | {', '.join(fixed_bits)}" if fixed_bits else ""
        axis.set_title(f"MCD vs B — {branch_name}{suffix}", fontsize=12.5,
                       fontweight="bold", pad=4)
        axis.tick_params(labelsize=10)
        axis.grid(alpha=0.18)

    energy_handles = [
        Line2D([0], [0], color="#333333", linestyle=energy_styles[value][0],
               marker=energy_styles[value][1], markerfacecolor="white",
               markersize=4, linewidth=1.25, label=f"{value:.3f} eV")
        for value in energy_values
    ]
    if energy_handles:
        energy_legend = axes[0].legend(
            handles=energy_handles, title="Processed energy", loc="upper left",
            fontsize=8.5, title_fontsize=9,
            ncol=2 if len(energy_handles) > 3 else 1,
            framealpha=0.88, borderpad=0.35, labelspacing=0.25,
            handlelength=2.1, columnspacing=0.8,
        )
        axes[0].add_artist(energy_legend)

    numeric_values = np.asarray([
        float(value) for value in (record_order_value(record, order_variable) for record in records)
        if isinstance(value, (int, float)) and np.isfinite(value)
    ], float)
    isolated = isolated_condition_values(numeric_values) if order_variable != "Source" else set()
    mapped = np.asarray([value for value in numeric_values if value not in isolated], float)
    if mapped.size:
        low, high = float(mapped.min()), float(mapped.max())
        if low < 0.0 < high:
            norm = TwoSlopeNorm(vmin=low, vcenter=0.0, vmax=high)
        elif high > low:
            norm = Normalize(vmin=low, vmax=high)
        else:
            norm = Normalize(vmin=low - 0.5, vmax=high + 0.5)
        cax = figure.add_axes([0.842, 0.14, 0.025, 0.76])
        colorbar = figure.colorbar(ScalarMappable(norm=norm, cmap=colormaps[palette]), cax=cax)
        descriptor = ORDER_VARIABLES.get(order_variable, (order_variable, "", None))
        colorbar.set_label(f"{descriptor[0]} ({descriptor[1]})" if descriptor[1] else descriptor[0],
                           fontsize=10)
        colorbar.ax.tick_params(labelsize=8.5)
    if isolated:
        isolated_text = ", ".join(f"{value:g}" for value in sorted(isolated))
        axes[0].legend(
            handles=[Patch(facecolor="#4a4a4a", label=f"Isolated: {isolated_text}")],
            loc="lower right", fontsize=8, framealpha=0.88, borderpad=0.3,
        )
    figure.savefig(path, dpi=figure.dpi, facecolor="white", edgecolor="none",
                   bbox_inches="tight", pad_inches=0.03)


def compact_slope_table(
    records: Sequence[ProcessedMcdRecord], comparison_variable: str
) -> pd.DataFrame:
    """One analysis-ready row per condition, with both branch slopes."""
    axis_labels = {
        "E-field": "E-field (V)", "Temperature": "Temperature (K)",
        "Doping": "Doping (V)", "Vtg": "Vtg (V)", "Vbg": "Vbg (V)",
        "Vbias": "Vbias (V)", "Energy": "Energy (eV)",
    }
    axis_label = axis_labels.get(comparison_variable, comparison_variable)
    rows = []
    for record in records:
        axis_value = (
            record.center_ev if comparison_variable == "Energy"
            else record.condition_value(comparison_variable)
        )
        rows.append({
            axis_label: axis_value,
            "Energy (eV)": record.center_ev,
            "Increasing slope (MCD/T)": record.increasing_slope_per_t,
            "Decreasing slope (MCD/T)": record.decreasing_slope_per_t,
        })
    return pd.DataFrame(rows)


def compact_conditions_table(
    records: Sequence[ProcessedMcdRecord], comparison_variable: str
) -> pd.DataFrame:
    rows = []
    used: dict[str, int] = {}
    for index, record in enumerate(records, start=1):
        prefix = _curve_label(record, comparison_variable, index)
        used[prefix] = used.get(prefix, 0) + 1
        if used[prefix] > 1:
            prefix = f"{prefix}_{used[prefix]}"
        rows.append({
            "Column prefix": prefix,
            "Source file": record.source_file,
            "Doping (V)": record.condition_value("Doping"),
            "E-field (V)": record.condition_value("E-field"),
            "Temperature setpoint (K)": record.temperature_setpoint_k or record.condition_value("T"),
            "Temperature measured (K)": record.temperature_measured_k,
            "Energy (eV)": record.center_ev,
            "Width (meV)": record.width_mev,
        })
    return pd.DataFrame(rows)


def _export_slope_png(
    path: Path, records: Sequence[ProcessedMcdRecord], comparison_variable: str
) -> None:
    from matplotlib.backends.backend_agg import FigureCanvasAgg
    from matplotlib.figure import Figure

    figure = Figure(figsize=(6.0, 4.5), dpi=300, facecolor="white")
    FigureCanvasAgg(figure)
    axis = figure.add_axes([0.16, 0.14, 0.80, 0.75])
    x = np.asarray([
        record.center_ev if comparison_variable == "Energy"
        else record.condition_value(comparison_variable)
        for record in records
    ], float)
    for values, label, linestyle, filled in (
        ([record.increasing_slope_per_t for record in records], "B increasing", "-", True),
        ([record.decreasing_slope_per_t for record in records], "B decreasing", "--", False),
    ):
        y = np.asarray([np.nan if value is None else value for value in values], float)
        valid = np.isfinite(x) & np.isfinite(y)
        if np.any(valid):
            axis.plot(
                x[valid], y[valid], linestyle=linestyle, marker="o", linewidth=1.8,
                markersize=5.5, markeredgewidth=1.0,
                color="#3568a8", markerfacecolor="#3568a8" if filled else "white",
                markeredgecolor="#3568a8", label=label,
            )
    axis.axhline(0.0, color="#555", linewidth=0.7)
    descriptor = ORDER_VARIABLES.get(comparison_variable, (comparison_variable, "", None))
    x_label = f"{descriptor[0]} ({descriptor[1]})" if descriptor[1] else descriptor[0]
    axis.set_xlabel(x_label, fontsize=12)
    axis.set_ylabel("Low-field MCD slope (MCD/T)", fontsize=12)

    fixed_conditions: list[str] = []
    condition_descriptors = (
        ("Doping", "D", "V"),
        ("E-field", "E-field", "V"),
        ("Temperature", "T", "K"),
        ("Vbias", "Vbias", "V"),
        ("Energy", "E", "eV"),
    )
    for variable, label, unit in condition_descriptors:
        if variable == comparison_variable:
            continue
        if variable == "Energy":
            values = [record.center_ev for record in records]
        elif variable == "Width":
            values = [record.width_mev for record in records]
        else:
            key = "T" if variable == "Temperature" else variable
            values = [record.condition_value(key) for record in records]
        finite = [float(value) for value in values if value is not None and np.isfinite(value)]
        if finite and len(finite) == len(records) and np.allclose(finite, finite[0], atol=1e-6, rtol=0):
            fixed_conditions.append(f"{label} = {finite[0]:g} {unit}")
    title = f"MCD slope vs {comparison_variable}"
    if fixed_conditions:
        title += "\n" + " | ".join(fixed_conditions)
    axis.set_title(title, fontsize=12.5, fontweight="bold", pad=4, linespacing=1.15)
    axis.tick_params(axis="both", labelsize=10.5)
    axis.grid(alpha=0.18)
    handles, labels = axis.get_legend_handles_labels()
    if handles:
        axis.legend(loc="best", fontsize=10, framealpha=0.9, borderpad=0.4)
    figure.savefig(path, dpi=figure.dpi, facecolor="white", edgecolor="none",
                   bbox_inches="tight", pad_inches=0.03)


def export_mcd_extract(
    records: Sequence[ProcessedMcdRecord],
    output_dir: str | Path,
    *,
    branches: Sequence[McdBranch] = BRANCHES,
    filters: McdExtractFilters | None = None,
    energy_tolerance_mev: float = 5.0,
    order_by: str = "Auto",
    descending: bool = False,
    palette: str = "viridis",
    export_csv: bool = False,
    series_groups: Sequence[McdSeries] | None = None,
) -> dict[str, Path]:
    """Export branch-only Origin data, compact summaries, plots, and settings."""
    from openpyxl import Workbook

    if not records:
        raise ValueError("Select at least one processed MCD(B) result to export.")
    if not branches:
        raise ValueError("Select at least one B-sweep branch to export.")
    effective_order = (
        series_groups[0].variable
        if order_by == "Auto" and series_groups and len(series_groups) == 1
        else order_by
    )
    ordered_records, resolved_order = order_mcd_records(
        records, effective_order, descending=descending
    )
    comparison_variable = resolved_order
    selected_palette = palette if palette in PALETTES else "viridis"
    colors = assign_plot_colors(ordered_records, selected_palette, comparison_variable)
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)
    base = _unused_export_base(
        out, descriptive_export_base(ordered_records, resolved_order, descending)
    )
    origin_workbook_path = out / f"{base}_Origin.xlsx"
    summary_workbook_path = out / f"{base}_Summary.xlsx"
    increasing_png = out / f"{base}_MCD_increasing_vs_B.png"
    decreasing_png = out / f"{base}_MCD_decreasing_vs_B.png"
    slope_png = out / f"{base}_Slope_vs_{resolved_order.replace('-', '')}.png"
    settings_path = out / f"{base}_settings.json"

    origin_workbook = Workbook()
    origin_workbook.remove(origin_workbook.active)
    branch_tables: dict[McdBranch, pd.DataFrame] = {}
    for branch in branches:
        table = origin_branch_table(ordered_records, branch, resolved_order)
        branch_tables[branch] = table
        _write_dataframe_sheet(
            origin_workbook, "Increasing" if branch == "B increasing" else "Decreasing", table
        )
    origin_workbook.properties.title = f"Origin-ready processed MCD(B), ordered by {resolved_order}"
    origin_workbook.properties.subject = "Independent XY pairs; import with repeating (XY) designation"
    origin_workbook.save(origin_workbook_path)

    summary_workbook = Workbook()
    summary_workbook.remove(summary_workbook.active)
    _write_dataframe_sheet(
        summary_workbook, "Slopes", compact_slope_table(ordered_records, comparison_variable)
    )
    _write_dataframe_sheet(
        summary_workbook, "Conditions", compact_conditions_table(ordered_records, comparison_variable)
    )
    summary_workbook.properties.title = f"Processed MCD summary ordered by {resolved_order}"
    summary_workbook.properties.subject = "Low-field slopes and essential condition metadata"
    summary_workbook.save(summary_workbook_path)
    if "B increasing" in branches:
        _export_comparison_png(
            increasing_png, ordered_records, ("B increasing",), colors,
            comparison_variable, selected_palette,
        )
    if "B decreasing" in branches:
        _export_comparison_png(
            decreasing_png, ordered_records, ("B decreasing",), colors,
            comparison_variable, selected_palette,
        )
    _export_slope_png(slope_png, ordered_records, comparison_variable)

    paths: dict[str, Path] = {
        "origin_xlsx": origin_workbook_path,
        "summary_xlsx": summary_workbook_path,
        "slope_png": slope_png,
        "settings": settings_path,
    }
    if "B increasing" in branches:
        paths["increasing_png"] = increasing_png
    if "B decreasing" in branches:
        paths["decreasing_png"] = decreasing_png
    if export_csv:
        for branch, table in branch_tables.items():
            key = "increasing_csv" if branch == "B increasing" else "decreasing_csv"
            csv_path = out / f"{base}_{'Increasing' if branch == 'B increasing' else 'Decreasing'}.csv"
            table.to_csv(csv_path, index=False)
            paths[key] = csv_path
    selection = {
        "schema_version": 2,
        "created_utc": datetime.now(timezone.utc).isoformat(),
        "branches": list(branches),
        "energy_group_tolerance_mev": float(energy_tolerance_mev),
        "filters": asdict(filters) if filters is not None else {},
        "plot": {
            "order_requested": order_by,
            "order_resolved": resolved_order,
            "descending": bool(descending),
            "palette": selected_palette,
            "colors": colors,
        },
        "series": [
            {
                "series_id": series.series_id,
                "variable": series.variable,
                "label": series.label,
                "record_ids": [record.record_id for record in series.records],
                "fixed_conditions": series.fixed_conditions,
            }
            for series in (series_groups or ())
        ],
        "records": [
            {
                "record_id": record.record_id,
                "source_file": record.source_file,
                "settings_path": str(record.settings_path),
                "trace_path": str(record.trace_path),
                "energy_ev": record.center_ev,
                "width_mev": record.width_mev,
                "conditions": {key: list(value) for key, value in record.acquisition_conditions.items()},
                "condition_sources": record.condition_sources,
            }
            for record in ordered_records
        ],
        "outputs": [path.name for path in paths.values()],
    }
    settings_path.write_text(json.dumps(selection, indent=2, sort_keys=True), encoding="utf-8")
    return paths
