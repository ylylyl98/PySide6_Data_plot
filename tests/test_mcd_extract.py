from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from core.mcd_extract import (
    McdExtractFilters,
    clear_mcd_trace_cache,
    discover_processed_mcd,
    energy_cluster_centers,
    export_mcd_extract,
    filter_processed_mcd,
    index_processed_mcd_settings,
    isolated_condition_values,
    load_branch_traces,
    mcd_catalog_database_path,
    newest_mcd_versions,
    organize_mcd_series,
    order_mcd_records,
    origin_branch_table,
    resolve_order_variable,
    slope_summary,
)


class McdExtractTests(unittest.TestCase):
    def test_isolated_condition_detection_uses_spacing_not_zero_value(self) -> None:
        self.assertEqual(isolated_condition_values([0, 20, 21, 22, 23]), {0.0})
        self.assertEqual(isolated_condition_values([5, 7, 9]), set())
        self.assertEqual(isolated_condition_values([-35, -34, -33, 0]), {0.0})
        self.assertEqual(isolated_condition_values([5, 7, 9, 50]), {50.0})

    def _write_result(
        self,
        root: Path,
        name: str,
        *,
        energy: float,
        doping: float,
        efield: float,
        increasing_slope: float = 2.0,
        decreasing_slope: float = 3.0,
        width: float = 5.0,
        temperature: float | None = 4.0,
        measurement_json_temperature: float | None = None,
        source_file: str | None = None,
        created_utc: str = "2026-08-25T12:00:00+00:00",
        vtg: float | None = None,
        vbg: float | None = None,
    ) -> Path:
        package = root / "Processed Data" / "MCD" / f"{name}_MCD"
        package.mkdir(parents=True, exist_ok=True)
        tag = f"E{energy:.6f}eV_W{width:g}meV"
        csv_name = f"{name}_MCD_vs_B_{tag}.csv"
        settings_name = f"{name}_MCD_settings_{tag}.json"
        increasing_b = np.array([-0.2, 0.0, 0.2])
        decreasing_b = np.array([0.2, 0.0, -0.2])
        pd.DataFrame({
            "B_increasing_T": increasing_b,
            "corrected_signed_mean_increasing": 0.1 + increasing_slope * increasing_b,
            "corrected_field_signed_absolute_mean_increasing": np.abs(increasing_b),
            "corrected_integral_increasing": 0.01 * increasing_b,
            "B_decreasing_T": decreasing_b,
            "corrected_signed_mean_decreasing": -0.2 + decreasing_slope * decreasing_b,
            "corrected_field_signed_absolute_mean_decreasing": np.abs(decreasing_b),
            "corrected_integral_decreasing": 0.02 * decreasing_b,
            "low_field_mcd_slope_increasing_per_T": increasing_slope,
            "low_field_mcd_slope_decreasing_per_T": decreasing_slope,
        }).to_csv(package / csv_name, index=False)
        payload = {
            "schema_version": 1,
            "workflow": "MCD",
            "source_file": source_file or f"{name}.csv",
            "package": package.name,
            "created_utc": created_utc,
            "outputs": [csv_name, settings_name],
            "mcd_b": {
                "center_ev": energy,
                "width_mev": width,
                "primary_metric": "mean",
                "fit_near_zero": True,
                "fit_window_t": 0.2,
                "low_field_mcd_slope_increasing_per_T": increasing_slope,
                "low_field_mcd_slope_decreasing_per_T": decreasing_slope,
            },
            "acquisition_conditions": {
                "Doping": [doping, doping],
                "E-field": [efield, efield],
            },
        }
        if temperature is not None:
            payload["acquisition_conditions"]["T"] = [temperature, temperature]
        if vtg is not None:
            payload["acquisition_conditions"]["Vtg"] = [vtg, vtg]
        if vbg is not None:
            payload["acquisition_conditions"]["Vbg"] = [vbg, vbg]
        (package / settings_name).write_text(json.dumps(payload), encoding="utf-8")
        if measurement_json_temperature is not None:
            (root / f"{name}.metadata.json").write_text(
                json.dumps({"measurement": {"sample_Tmid_K": measurement_json_temperature}}),
                encoding="utf-8",
            )
            source_stem = Path(source_file or f"{name}.csv").stem
            (root / f"{source_stem}.metadata.json").write_text(
                json.dumps({"measurement": {"sample_Tmid_K": measurement_json_temperature}}),
                encoding="utf-8",
            )
        return package

    def test_catalog_and_tolerant_filters_use_json_metadata(self) -> None:
        with tempfile.TemporaryDirectory() as folder_text:
            root = Path(folder_text)
            self._write_result(root, "a", energy=1.650, doping=6.300, efield=0.2)
            self._write_result(root, "b", energy=1.654, doping=6.306, efield=0.2)
            self._write_result(root, "c", energy=1.700, doping=7.0, efield=0.3)
            records = discover_processed_mcd(root)
            self.assertEqual(len(records), 3)
            filtered = filter_processed_mcd(
                records,
                McdExtractFilters(
                    doping_v=6.3,
                    doping_tolerance_v=0.01,
                    efield_v=0.2,
                    efield_tolerance_v=0.001,
                    temperature_k=4.05,
                    temperature_tolerance_k=0.1,
                    energy_min_ev=1.64,
                    energy_max_ev=1.66,
                ),
            )
            self.assertEqual([record.source_file for record in filtered], ["a.csv", "b.csv"])
            groups = energy_cluster_centers(filtered, 5.0)
            self.assertEqual(len(set(groups.values())), 1)

    def test_temperature_falls_back_to_measurement_json_with_provenance(self) -> None:
        with tempfile.TemporaryDirectory() as folder_text:
            root = Path(folder_text)
            self._write_result(
                root, "json_temp", energy=1.65, doping=1.0, efield=0.0,
                temperature=None, measurement_json_temperature=77.0,
            )
            record = discover_processed_mcd(root)[0]
            self.assertEqual(record.condition_value("T"), 77.0)
            self.assertIn("measurement JSON", record.condition_sources["T"])

    def test_temperature_filename_precedes_missing_metadata_default(self) -> None:
        with tempfile.TemporaryDirectory() as folder_text:
            root = Path(folder_text)
            self._write_result(
                root, "encoded_temp", energy=1.65, doping=1.0, efield=0.0,
                temperature=None, source_file="sample_2p5K_run.csv",
            )
            record = discover_processed_mcd(root)[0]
            self.assertEqual(record.condition_value("T"), 2.5)
            self.assertEqual(record.condition_sources["T"], "source filename setpoint")

    def test_temperature_setpoint_is_separate_from_measured_temperature(self) -> None:
        with tempfile.TemporaryDirectory() as folder_text:
            root = Path(folder_text)
            self._write_result(
                root, "setpoint_temp", energy=1.65, doping=1.0, efield=0.0,
                temperature=None, source_file="sample_2p5K_run.csv",
                measurement_json_temperature=2.73,
            )
            record = discover_processed_mcd(root)[0]
            self.assertEqual(record.condition_value("T"), 2.5)
            self.assertEqual(record.temperature_setpoint_k, 2.5)
            self.assertEqual(record.temperature_measured_k, 2.73)
            self.assertEqual(record.condition_sources["T"], "source filename setpoint")

    def test_missing_temperature_is_explicitly_assumed_to_be_1p67_k(self) -> None:
        with tempfile.TemporaryDirectory() as folder_text:
            root = Path(folder_text)
            self._write_result(
                root, "missing_temp", energy=1.65, doping=1.0, efield=0.0,
                temperature=None,
            )
            record = discover_processed_mcd(root)[0]
            self.assertEqual(record.condition_value("T"), 1.67)
            self.assertIn("assumed default", record.condition_sources["T"])

    def test_auto_order_prefers_efield_then_temperature(self) -> None:
        with tempfile.TemporaryDirectory() as folder_text:
            root = Path(folder_text)
            self._write_result(root, "f_high", energy=1.65, doping=2.0, efield=0.3)
            self._write_result(root, "f_low", energy=1.65, doping=2.0, efield=-0.1)
            field_records = discover_processed_mcd(root)
            ordered, variable = order_mcd_records(field_records)
            self.assertEqual(variable, "E-field")
            self.assertEqual([record.condition_value("E-field") for record in ordered], [-0.1, 0.3])

        with tempfile.TemporaryDirectory() as folder_text:
            root = Path(folder_text)
            self._write_result(root, "t_high", energy=1.65, doping=2.0, efield=0.0, temperature=80)
            self._write_result(root, "t_low", energy=1.65, doping=2.0, efield=0.0, temperature=4)
            temperature_records = discover_processed_mcd(root)
            self.assertEqual(resolve_order_variable(temperature_records), "Temperature")
            ordered, variable = order_mcd_records(temperature_records, descending=True)
            self.assertEqual(variable, "Temperature")
            self.assertEqual([record.condition_value("T") for record in ordered], [80.0, 4.0])

    def test_branch_traces_and_slopes_remain_distinct(self) -> None:
        with tempfile.TemporaryDirectory() as folder_text:
            root = Path(folder_text)
            self._write_result(
                root, "branch", energy=1.65, doping=1.0, efield=0.0,
                increasing_slope=2.0, decreasing_slope=3.0,
            )
            record = discover_processed_mcd(root)[0]
            traces = load_branch_traces(record)
            self.assertEqual(set(traces["branch"]), {"B increasing", "B decreasing"})
            self.assertEqual(len(traces), 6)
            slopes = slope_summary([record])
            self.assertEqual(slopes["branch"].tolist(), ["B increasing", "B decreasing"])
            self.assertTrue(np.allclose(slopes["slope_per_T"], [2.0, 3.0]))
            self.assertTrue(np.allclose(slopes["intercept"], [0.1, -0.2]))
            self.assertTrue(np.allclose(slopes["fit_r_squared"], [1.0, 1.0]))

    def test_trace_table_is_cached_between_preview_and_export_reads(self) -> None:
        with tempfile.TemporaryDirectory() as folder_text:
            root = Path(folder_text)
            self._write_result(root, "cached", energy=1.65, doping=1.0, efield=0.0)
            record = discover_processed_mcd(root)[0]
            clear_mcd_trace_cache()
            real_read_csv = pd.read_csv
            with patch("core.mcd_extract.pd.read_csv", wraps=real_read_csv) as read_csv:
                load_branch_traces(record)
                load_branch_traces(record, ("B increasing",))
            self.assertEqual(read_csv.call_count, 1)

    def test_catalog_is_incremental_and_keeps_per_condition_energies(self) -> None:
        with tempfile.TemporaryDirectory() as folder_text:
            root = Path(folder_text)
            self._write_result(root, "f_low", energy=1.650, doping=2.0, efield=-0.2)
            self._write_result(root, "f_mid", energy=1.674, doping=2.0, efield=0.0)
            self._write_result(root, "f_high", energy=1.701, doping=2.0, efield=0.2)
            first = discover_processed_mcd(root)
            catalog = mcd_catalog_database_path(root)
            self.assertTrue(catalog.is_file())
            with patch.object(Path, "rglob", side_effect=AssertionError("unexpected rescan")):
                second = discover_processed_mcd(root)
            self.assertEqual([record.record_id for record in second], [record.record_id for record in first])
            series = organize_mcd_series(second)
            self.assertEqual(len(series), 1)
            self.assertEqual(series[0].variable, "E-field")
            self.assertEqual(
                [record.center_ev for record in series[0].records],
                [1.650, 1.674, 1.701],
            )
            self.assertEqual(
                organize_mcd_series(
                    second, "Temperature", include_singletons=False
                ),
                [],
            )

    def test_newest_version_removes_only_matching_reprocessing(self) -> None:
        with tempfile.TemporaryDirectory() as folder_text:
            root = Path(folder_text)
            self._write_result(
                root, "old", energy=1.65, doping=2.0, efield=0.0,
                source_file="same.csv", created_utc="2026-08-24T12:00:00+00:00",
            )
            self._write_result(
                root, "new", energy=1.65, doping=2.0, efield=0.0,
                source_file="same.csv", created_utc="2026-08-25T12:00:00+00:00",
            )
            self._write_result(
                root, "other_energy", energy=1.70, doping=2.0, efield=0.0,
                source_file="same.csv", created_utc="2026-08-23T12:00:00+00:00",
            )
            newest, older = newest_mcd_versions(discover_processed_mcd(root))
            self.assertEqual({record.center_ev for record in newest}, {1.65, 1.70})
            self.assertEqual([record.package for record in older], ["old_MCD"])

    def test_new_result_upsert_appears_without_directory_rescan(self) -> None:
        with tempfile.TemporaryDirectory() as folder_text:
            root = Path(folder_text)
            self._write_result(root, "first", energy=1.65, doping=6.3, efield=0.0)
            self.assertEqual(len(discover_processed_mcd(root)), 1)

            second_package = self._write_result(
                root, "second", energy=1.68, doping=6.3, efield=20.0
            )
            # A manually copied result is intentionally invisible until it is
            # indexed or the user requests a full rebuild.
            self.assertEqual(len(discover_processed_mcd(root)), 1)
            second_settings = next(second_package.glob("*_MCD_settings*.json"))
            self.assertTrue(index_processed_mcd_settings(second_settings))
            with patch.object(Path, "rglob", side_effect=AssertionError("unexpected rescan")):
                records = discover_processed_mcd(root)
            self.assertEqual({record.source_file for record in records}, {"first.csv", "second.csv"})

    def test_rebuild_reconciles_results_copied_outside_the_app(self) -> None:
        with tempfile.TemporaryDirectory() as folder_text:
            root = Path(folder_text)
            self._write_result(root, "first", energy=1.65, doping=6.3, efield=0.0)
            self.assertEqual(len(discover_processed_mcd(root)), 1)
            self._write_result(root, "copied", energy=1.70, doping=6.3, efield=25.0)
            self.assertEqual(len(discover_processed_mcd(root)), 1)
            self.assertEqual(len(discover_processed_mcd(root, rebuild_catalog=True)), 2)

    def test_efield_series_allows_dual_gate_voltages_to_change(self) -> None:
        with tempfile.TemporaryDirectory() as folder_text:
            root = Path(folder_text)
            self._write_result(
                root, "field_0", energy=1.65, doping=6.3, efield=0.0,
                vtg=3.15, vbg=2.90,
            )
            self._write_result(
                root, "field_20", energy=1.68, doping=6.3, efield=20.0,
                vtg=13.15, vbg=-6.30,
            )
            records = discover_processed_mcd(root)
            series = organize_mcd_series(
                records, "E-field", include_singletons=False
            )
            self.assertEqual(len(series), 1)
            self.assertEqual(
                [record.condition_value("E-field") for record in series[0].records],
                [0.0, 20.0],
            )
            self.assertNotIn("Vtg", series[0].fixed_conditions)
            self.assertNotIn("Vbg", series[0].fixed_conditions)

    def test_export_writes_origin_workbook_png_and_settings_with_descriptive_base(self) -> None:
        with tempfile.TemporaryDirectory() as folder_text:
            root = Path(folder_text)
            self._write_result(root, "one", energy=1.650, doping=2.0, efield=0.1)
            self._write_result(root, "two", energy=1.653, doping=2.0, efield=0.1)
            records = discover_processed_mcd(root)
            paths = export_mcd_extract(records, root / "extracts", energy_tolerance_mev=5.0)
            self.assertEqual(
                set(paths),
                {"origin_xlsx", "summary_xlsx", "increasing_png", "decreasing_png", "slope_png", "settings"},
            )
            self.assertIn("D2V", paths["origin_xlsx"].name)
            self.assertIn("F0.1V", paths["origin_xlsx"].name)
            self.assertIn("byEnergy_asc", paths["origin_xlsx"].name)
            self.assertTrue(paths["origin_xlsx"].name.endswith("_Origin.xlsx"))
            self.assertTrue(paths["summary_xlsx"].name.endswith("_Summary.xlsx"))
            workbook = load_workbook(paths["origin_xlsx"], data_only=False)
            self.assertEqual(workbook.sheetnames, ["Increasing", "Decreasing"])
            increasing_headers = [cell.value for cell in workbook["Increasing"][1]]
            decreasing_headers = [cell.value for cell in workbook["Decreasing"][1]]
            self.assertTrue(any(str(value).endswith("_B_T") for value in increasing_headers))
            self.assertTrue(any(str(value).endswith("_MCD") for value in increasing_headers))
            self.assertFalse(any("_fit_" in str(value) or "_slope_" in str(value) for value in increasing_headers))
            self.assertTrue(any(str(value).endswith("_B_T") for value in decreasing_headers))
            self.assertTrue(any("E1.65eV" in str(value) for value in increasing_headers))
            summary = load_workbook(paths["summary_xlsx"], data_only=False)
            self.assertEqual(summary.sheetnames, ["Slopes", "Conditions"])
            slope_headers = [cell.value for cell in summary["Slopes"][1]]
            self.assertIn("Increasing slope (MCD/T)", slope_headers)
            self.assertIn("Decreasing slope (MCD/T)", slope_headers)
            self.assertEqual(summary["Slopes"]["A1"].fill.fill_type, None)
            manifest = json.loads(paths["settings"].read_text(encoding="utf-8"))
            self.assertEqual(manifest["plot"]["order_resolved"], "Energy")
            self.assertEqual(manifest["plot"]["palette"], "viridis")
            self.assertFalse(list((root / "extracts").glob("*.csv")))
            self.assertTrue(all(path.is_file() for path in paths.values()))

    def test_optional_csv_export_separates_branches(self) -> None:
        with tempfile.TemporaryDirectory() as folder_text:
            root = Path(folder_text)
            self._write_result(root, "one", energy=1.65, doping=1.0, efield=0.0)
            record = discover_processed_mcd(root)[0]
            paths = export_mcd_extract([record], root / "extracts", export_csv=True)
            self.assertIn("increasing_csv", paths)
            self.assertIn("decreasing_csv", paths)
            self.assertIn("_Increasing.csv", paths["increasing_csv"].name)
            self.assertIn("_Decreasing.csv", paths["decreasing_csv"].name)
            table = origin_branch_table([record], "B increasing", "Source")
            self.assertTrue(any(column.endswith("_B_T") for column in table.columns))
            self.assertTrue(any(column.endswith("_MCD") for column in table.columns))
            self.assertFalse(any("_fit_" in column or "_slope_" in column for column in table.columns))

    def test_grouped_export_adds_compact_conditions_and_preserves_energy(self) -> None:
        with tempfile.TemporaryDirectory() as folder_text:
            root = Path(folder_text)
            self._write_result(root, "low", energy=1.65, doping=2.0, efield=-0.2)
            self._write_result(root, "high", energy=1.71, doping=2.0, efield=0.2)
            records = discover_processed_mcd(root)
            series = organize_mcd_series(records)
            paths = export_mcd_extract(
                records, root / "extracts", series_groups=series
            )
            workbook = load_workbook(paths["summary_xlsx"], data_only=True)
            self.assertEqual(workbook.sheetnames, ["Slopes", "Conditions"])
            headers = [cell.value for cell in workbook["Conditions"][1]]
            energy_column = headers.index("Energy (eV)") + 1
            energies = {
                workbook["Conditions"].cell(row, energy_column).value
                for row in range(2, workbook["Conditions"].max_row + 1)
            }
            self.assertEqual(energies, {1.65, 1.71})


if __name__ == "__main__":
    unittest.main()
