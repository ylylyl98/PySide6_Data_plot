from __future__ import annotations

import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import numpy as np

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from core import data_io
from core import loader


def _write_map_xlsx(path: Path, header: list, rows: list) -> None:
    """Write a small precomputed-map workbook mirroring the dR_R layout."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "dR_R"
    ws.append(header)
    for row in rows:
        ws.append(row)
    wb.save(path)


class XlsxMapInputTests(unittest.TestCase):
    def setUp(self) -> None:
        for fn in (loader._load_xlsx_map_cached,):
            clear = getattr(fn, "cache_clear", None)
            if callable(clear):
                clear()

    def _make_folder(self) -> Path:
        self._tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp.cleanup)
        return Path(self._tmp.name)

    def test_is_xlsx_map_file(self) -> None:
        self.assertTrue(loader.is_xlsx_map_file("dR_R.xlsx"))
        self.assertTrue(loader.is_xlsx_map_file("dR_R.XLSX"))
        self.assertFalse(loader.is_xlsx_map_file("dR_R.csv"))

    def test_resolve_xlsx_y_label_default_and_options(self) -> None:
        self.assertEqual(loader.XLSX_Y_LABEL_OPTIONS, ("Doping (V)", "Efield (V)"))
        self.assertEqual(loader.resolve_xlsx_y_label("auto"), "Doping (V)")
        self.assertEqual(loader.resolve_xlsx_y_label("doping"), "Doping (V)")
        self.assertEqual(loader.resolve_xlsx_y_label("efield"), "Efield (V)")
        self.assertEqual(loader.resolve_xlsx_y_label("Efield (V)"), "Efield (V)")

    def test_load_xlsx_map_transposes_and_sorts(self) -> None:
        folder = self._make_folder()
        path = folder / "dR_R.xlsx"
        # Corner label duplicates the numeric 0.0 doping header; this must survive.
        header = ["0", 0.0, -1.0, 1.0]
        rows = [
            [1.8, 10.0, 11.0, 12.0],
            [1.7, 20.0, 21.0, 22.0],
        ]
        _write_map_xlsx(path, header, rows)

        cube = loader.load_xlsx_map(str(folder), "dR_R.xlsx")

        np.testing.assert_array_equal(cube.energy, np.array([1.7, 1.8]))
        np.testing.assert_array_equal(cube.gate, np.array([0.0, -1.0, 1.0]))
        # Z is transposed to (doping, energy) and rows reordered with ascending energy.
        np.testing.assert_array_equal(
            cube.Z,
            np.array(
                [
                    [20.0, 10.0],
                    [21.0, 11.0],
                    [22.0, 12.0],
                ]
            ),
        )
        self.assertEqual(cube.gate_label, "Doping (V)")
        self.assertEqual(cube.cbar_label, "dR/R")
        self.assertEqual(cube.title, "dR_R.xlsx")

    def test_load_xlsx_map_y_label_choice_does_not_change_arrays(self) -> None:
        folder = self._make_folder()
        path = folder / "dR_R.xlsx"
        _write_map_xlsx(path, ["0", 0.0, 1.0], [[1.8, 10.0, 12.0], [1.7, 20.0, 22.0]])

        doping = loader.load_xlsx_map(str(folder), "dR_R.xlsx", y_label="Doping (V)")
        efield = loader.load_xlsx_map(str(folder), "dR_R.xlsx", y_label="Efield (V)")

        self.assertEqual(doping.gate_label, "Doping (V)")
        self.assertEqual(efield.gate_label, "Efield (V)")
        np.testing.assert_array_equal(doping.energy, efield.energy)
        np.testing.assert_array_equal(doping.gate, efield.gate)
        np.testing.assert_array_equal(doping.Z, efield.Z)

    def test_load_xlsx_map_single_sheet_fallback(self) -> None:
        folder = self._make_folder()
        path = folder / "map.xlsx"
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "OnlySheet"
        ws.append(["0", 0.0])
        ws.append([1.7, 20.0])
        wb.save(path)

        cube = loader.load_xlsx_map(str(folder), "map.xlsx")
        self.assertEqual(cube.title, "map.xlsx")

    def test_load_xlsx_map_rejects_multiple_sheets(self) -> None:
        folder = self._make_folder()
        path = folder / "map.xlsx"
        from openpyxl import Workbook

        wb = Workbook()
        wb.active.title = "SheetA"
        wb.create_sheet("SheetB")
        wb.save(path)

        with self.assertRaises(ValueError):
            loader.load_xlsx_map(str(folder), "map.xlsx")

    def test_load_xlsx_map_rejects_formula(self) -> None:
        folder = self._make_folder()
        path = folder / "map.xlsx"
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "dR_R"
        ws.append(["0", 0.0])
        ws.append([1.7, "=SUM(A1:B1)"])
        wb.save(path)

        with self.assertRaises(ValueError):
            loader.load_xlsx_map(str(folder), "map.xlsx")

    def test_load_xlsx_map_rejects_non_numeric(self) -> None:
        folder = self._make_folder()
        path = folder / "map.xlsx"
        _write_map_xlsx(path, ["0", 0.0], [[1.7, "oops"]])

        with self.assertRaises(ValueError):
            loader.load_xlsx_map(str(folder), "map.xlsx")

    def test_load_xlsx_map_rejects_duplicate_y(self) -> None:
        folder = self._make_folder()
        path = folder / "map.xlsx"
        _write_map_xlsx(path, ["0", 0.0, 0.0], [[1.7, 1.0, 2.0]])

        with self.assertRaises(ValueError):
            loader.load_xlsx_map(str(folder), "map.xlsx")

    def test_load_xlsx_map_rejects_ragged_row(self) -> None:
        folder = self._make_folder()
        path = folder / "map.xlsx"
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "dR_R"
        ws.append(["0", 0.0, 1.0])
        ws.append([1.7, 20.0])  # missing third column
        wb.save(path)

        with self.assertRaises(ValueError):
            loader.load_xlsx_map(str(folder), "map.xlsx")

    def test_load_xlsx_map_requires_root_file(self) -> None:
        folder = self._make_folder()
        with self.assertRaises(FileNotFoundError):
            loader.load_xlsx_map(str(folder), "missing.xlsx")

    def test_load_xlsx_map_supports_nested_and_absolute_sources(self) -> None:
        folder = self._make_folder()
        nested = folder / "Initial Data" / "old session"
        nested.mkdir(parents=True)
        path = nested / "dR_R.xlsx"
        _write_map_xlsx(path, ["0", -1.0, 1.0], [[1.7, 2.0, 3.0]])

        relative = loader.load_xlsx_map(str(folder), "Initial Data/old session/dR_R.xlsx")
        absolute = loader.load_xlsx_map(str(folder), str(path))

        self.assertTrue(np.array_equal(relative.Z, absolute.Z))

    def test_load_xlsx_map_rejects_non_xlsx_suffix(self) -> None:
        folder = self._make_folder()
        (folder / "dR_R.csv").write_text("a,b\n1,2\n")
        with self.assertRaises(ValueError):
            loader.load_xlsx_map(str(folder), "dR_R.csv")

    def test_list_map_input_files_includes_xlsx_and_csv(self) -> None:
        folder = self._make_folder()
        (folder / "run_10.csv").write_text("x")
        (folder / "run_2.csv").write_text("x")
        (folder / "map.xlsx").write_bytes(b"x")
        (folder / "note.txt").write_text("ignored")

        self.assertEqual(
            data_io.list_map_input_files(str(folder)),
            ["map.xlsx", "run_2.csv", "run_10.csv"],
        )
        # CSV-only listing is unchanged and does not include the workbook.
        self.assertEqual(
            data_io.list_csv_files(str(folder)),
            ["run_2.csv", "run_10.csv"],
        )

    def test_load_pl_dispatch_uses_xlsx_map_path(self) -> None:
        folder = self._make_folder()
        path = folder / "dR_R.xlsx"
        _write_map_xlsx(path, ["0", 0.0], [[1.7, 20.0]])

        cube = data_io.load_pl_cube(str(folder), "dR_R.xlsx")
        self.assertEqual(cube.gate_label, "Doping (V)")
        self.assertEqual(cube.cbar_label, "dR/R")
        np.testing.assert_array_equal(cube.gate, np.array([0.0]))


    def test_load_drr_map_cube_contract(self) -> None:
        folder = self._make_folder()
        path = folder / "dR_R.xlsx"
        _write_map_xlsx(path, ["0", 0.0, 1.0], [[1.8, 10.0, 12.0], [1.7, 20.0, 22.0]])

        default = data_io.load_drr_map_cube(str(folder), "dR_R.xlsx")
        efield = data_io.load_drr_map_cube(str(folder), "dR_R.xlsx", y_axis="efield")

        self.assertEqual(default.cbar_label, "dR/R")
        self.assertEqual(default.gate_label, "Doping (V)")
        self.assertEqual(efield.gate_label, "Efield (V)")
        np.testing.assert_array_equal(default.energy, efield.energy)
        np.testing.assert_array_equal(default.gate, efield.gate)
        np.testing.assert_array_equal(default.Z, efield.Z)


class XlsxMapDrrUiTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        from PySide6.QtWidgets import QApplication

        cls.app = QApplication.instance() or QApplication([])

    def setUp(self) -> None:
        from ui_qt.main_window import MainWindow

        with patch.object(MainWindow, "_restore_last_folder", lambda _self: None):
            self.window = MainWindow()

    def tearDown(self) -> None:
        self.window.close()

    def _combo_items(self):
        combo = self.window.drr_yaxis_combo
        return [combo.itemText(i) for i in range(combo.count())]

    def test_drr_yaxis_repopulates_for_xlsx_and_csv(self) -> None:
        combo = self.window.drr_yaxis_combo
        self.assertEqual(
            self._combo_items(),
            ["Auto / Default", "TG", "BG", "Bias", "Advanced..."],
        )

        self.window.drr_selected_files = ["dR_R.xlsx"]
        self.window._update_drr_selection_labels()
        self.assertEqual(
            self._combo_items(),
            ["Auto / Default", "Doping (V)", "Efield (V)"],
        )
        self.assertEqual(combo.currentText(), "Auto / Default")

        combo.blockSignals(True)
        try:
            combo.setCurrentText("Efield (V)")
        finally:
            combo.blockSignals(False)
        self.assertEqual(self.window._selected_y_axis_spec("drr"), "efield")

        self.window.drr_selected_files = ["run.csv"]
        self.window._update_drr_selection_labels()
        self.assertEqual(
            self._combo_items(),
            ["Auto / Default", "TG", "BG", "Bias", "Advanced..."],
        )

    def test_reject_mixed_xlsx_selection(self) -> None:
        self.window._reject_mixed_xlsx_selection([])
        self.window._reject_mixed_xlsx_selection(["a.csv", "b.csv"])
        self.window._reject_mixed_xlsx_selection(["dR_R.xlsx"])
        with self.assertRaises(ValueError):
            self.window._reject_mixed_xlsx_selection(["a.csv", "dR_R.xlsx"])
        with self.assertRaises(ValueError):
            self.window._reject_mixed_xlsx_selection(["a.xlsx", "b.xlsx"])


if __name__ == "__main__":
    unittest.main()
